from flask import Blueprint, render_template, request, redirect, url_for, flash, current_app, send_from_directory
from flask_login import login_user, logout_user, login_required, current_user
from werkzeug.utils import secure_filename
import os
from models import User, Todo, Director, Customer, Transaction, PettyCash, PettyCashCategory, Bank, BankTransaction, Installment, CustomerInstallment, Party, PartyLedger, Voucher, PartyCategory, ContraEntry, Employee, Attendance, Leave, Salary
from database import db
from logic import sync_to_excel, restore_from_excel, recalculate_party_ledger_balances, export_party_ledger_to_excel, generate_voucher_number, process_voucher_financials, get_daily_cash_report, get_due_payments_report, generate_contra_number, process_contra_financials
import pandas as pd
import io
from datetime import datetime
from flask import send_file, jsonify
import random
import string
import json
import calendar
from telegram_utils import send_telegram_message, send_telegram_document

import threading

def run_in_background(target, app, *args, **kwargs):
    """Universal helper to run a function in a background thread."""
    def run_with_context():
        with app.app_context():
            target(*args, **kwargs)
            
    thread = threading.Thread(target=run_with_context)
    thread.daemon = True
    thread.start()

_global_sync_lock = threading.Lock()

def get_director_initials(name):
    """Generates initials from director name, ignoring special characters."""
    if not name:
        return "DIR"
    
    # Filter for alphanumeric and groups
    clean_name = ''.join(c if c.isalnum() or c.isspace() else ' ' for c in name)
    parts = clean_name.split()
    
    if len(parts) >= 2:
        return (parts[0][0] + parts[1][0]).upper()
    return parts[0][:2].upper() if parts else "DI"

def generate_next_customer_id(director_id):
    """Generates next customer ID like AI-01, AH-01 etc."""
    director = Director.query.get(director_id)
    if not director:
        return "CUST-01"
    
    initials = get_director_initials(director.name)
    count = Customer.query.filter_by(director_id=director_id).count()
    return f"{initials}-{count + 1:02d}"

def recalculate_customer_totals(customer):
    """
    Recalculates a customer's total_price, total_paid and due_amount based on
    actual transactions and created installments.
    """
    # 1. Total Paid = Sum of all transactions
    customer.total_paid = sum(t.amount for t in customer.transactions)
    
    # 2. Sync installment milestones with current shares
    #    Iterate through all installments and update total_amount based on current shares
    for ci in customer.installments:
        if ci.installment:
            ci.total_amount = customer.shares * ci.installment.amount_per_share
            ci.due_amount = ci.total_amount - ci.paid_amount

    # 3. Total Expected = Sum of all CustomerInstallments
    total_expected = sum(ci.total_amount for ci in customer.installments)
    
    # 4. Sync total_price to reflect true installment total
    if customer.installments:
        customer.total_price = total_expected
    
    # 5. Due = Expected - Paid
    customer.due_amount = total_expected - customer.total_paid
    
    # 5. Sync Director Totals: use total_share (not assigned shares) × installment rates
    director = customer.director
    if director:
        director.total_paid = sum(c.total_paid for c in director.customers)
        # Total expected for director = total_share × sum of all installment amount_per_share
        from models import Installment
        total_rate_per_share = sum(inst.amount_per_share for inst in Installment.query.all())
        director_total_expected = director.total_share * total_rate_per_share
        director.total_due = director_total_expected - director.total_paid
    
    return customer

def background_sync_all(action_name="Data Update"):
    """
    Consolidates all high-latency tasks into a single background thread.
    This prevents multiple threads from competing for SQLite locks or CPU.
    """
    app = current_app._get_current_object()
    
    if not _global_sync_lock.acquire(blocking=False):
        from telegram_utils import log_debug
        log_debug(f"[{action_name}] Sync skipped: A sync is already in progress.")
        return
        
    def sync_task(app_obj):
        try:
            from sync_manager import sync_manager
            from telegram_utils import log_debug
            log_debug(f"Starting Background Tasks due to: {action_name}")
            
            sync_manager.sync_to_sheets()
            
            # 2. Master Excel Sync (Disabled/Moved elsewhere if needed to save time, or we can leave it)
            try:
                sync_to_excel()
            except Exception as e:
                log_debug(f"Background Excel sync failed: {e}")

            # 3. DB Backup with Cleanup
            try:
                from logic import create_db_backup
                create_db_backup()
                
                # Also send to Telegram (existing logic)
                db_path = app_obj.config.get('DATABASE_PATH')
                if db_path and os.path.exists(db_path):
                    from telegram_utils import send_telegram_document
                    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    caption = f"DB Backup triggered by: {action_name}\nTime: {timestamp}"
                    send_telegram_document(db_path, caption=caption)
            except Exception as e:
                log_debug(f"Background DB backup failed: {e}")

            # 4. Task Reminders
            try:
                from models import Todo
                from datetime import datetime, timedelta
                now = datetime.now()
                # Check for todos due today that haven't been reminded yet
                today_str = now.strftime('%Y-%m-%d')
                upcoming_todos = Todo.query.filter(
                    Todo.due_date == today_str,
                    Todo.is_completed == False,
                    Todo.reminder_sent == False
                ).all()
                
                if upcoming_todos:
                    from telegram_utils import send_telegram_message
                    for todo in upcoming_todos:
                        msg = f"🔔 *REMINDER*: {todo.task}\nCategory: {todo.category}\nDue Today"
                        if todo.due_time:
                            msg += f" at {todo.due_time}"
                        
                        send_telegram_message(msg)
                        todo.reminder_sent = True
                    db.session.commit()
            except Exception as e:
                log_debug(f"Background Task Reminders failed: {e}")
                
        except Exception as e:
             from telegram_utils import log_debug
             log_debug(f"Background Task Error: {e}")
        finally:
             _global_sync_lock.release()

    # Dispatch to thread
    run_in_background(sync_task, app, app)
# For backward compatibility with existing calls, we redefine these to call the consolidated worker
def trigger_sync():
    background_sync_all("Google Sheets Sync")

def trigger_excel_sync():
    background_sync_all("Excel Sync")

def backup_to_telegram(action_name="Database Update"):
    background_sync_all(action_name)

main = Blueprint('main', __name__)

@main.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(current_app.config['UPLOAD_FOLDER'], filename)

@main.route('/')
@login_required
def index():
    try:
        directors = Director.query.all()
        customers = Customer.query.all()
        grand_total_payable = sum(c.total_price for c in customers)
        grand_total_paid = sum(c.total_paid for c in customers)
        grand_total_due = sum(c.due_amount for c in customers)
        grand_total_outstanding = grand_total_payable - grand_total_paid
            
        total_bank_balance = sum(tx.credit - tx.debit for tx in BankTransaction.query.all())
        
        cash_income = sum(pc.amount for pc in PettyCash.query.filter_by(type='Income').all())
        cash_expense = sum(pc.amount for pc in PettyCash.query.filter_by(type='Expense').all())
        cash_in_hand = cash_income - cash_expense

        # Store today_todos safely
        from datetime import datetime
        today_str = datetime.now().strftime('%Y-%m-%d')
        today_todos = []
        try:
            today_todos = Todo.query.filter_by(
                user_id=current_user.id,
                due_date=today_str,
                is_completed=False
            ).all()
        except Exception as todo_err:
            print(f"Error fetching todos on index: {todo_err}")

        # Chart Data (Last 6 Months)
        from collections import defaultdict
        income_by_month = defaultdict(float)
        expense_by_month = defaultdict(float)
        
        def get_month_key(date_str):
            if not date_str: return None
            for fmt in ('%d-%m-%Y', '%Y-%m-%d'):
                try:
                    return datetime.strptime(date_str, fmt).strftime('%Y-%m')
                except ValueError:
                    pass
            return None

        # 1. Customer Transactions (Income)
        for tx in Transaction.query.all():
            m = get_month_key(tx.date)
            if m: income_by_month[m] += tx.amount
            
        # 2. Bank Transactions
        for tx in BankTransaction.query.all():
            m = get_month_key(tx.date)
            if m:
                income_by_month[m] += tx.credit
                expense_by_month[m] += tx.debit
                
        # 3. Petty Cash
        for pc in PettyCash.query.all():
            m = get_month_key(pc.date)
            if m:
                if pc.type == 'Income': income_by_month[m] += pc.amount
                else: expense_by_month[m] += pc.amount

        # Prepare labels and values for the last 6 months
        import calendar
        from datetime import date
        chart_labels = []
        chart_income = []
        chart_expense = []
        
        current_date = date.today().replace(day=1)
        for _ in range(6):
            m_key = current_date.strftime('%Y-%m')
            chart_labels.insert(0, calendar.month_name[current_date.month][:3] + " " + str(current_date.year)[2:])
            chart_income.insert(0, income_by_month[m_key])
            chart_expense.insert(0, expense_by_month[m_key])
            # Move to previous month
            if current_date.month == 1:
                current_date = current_date.replace(year=current_date.year - 1, month=12)
            else:
                current_date = current_date.replace(month=current_date.month - 1)

        parties = Party.query.all()
        total_party_due = sum(p.current_balance for p in parties if p.current_balance > 0)
        total_party_advance = abs(sum(p.current_balance for p in parties if p.current_balance < 0))

        return render_template('index.html', directors=directors, 
                             grand_total_payable=grand_total_payable, 
                             grand_total_paid=grand_total_paid, 
                             grand_total_due=grand_total_due,
                             grand_total_outstanding=grand_total_outstanding,
                             total_bank_balance=total_bank_balance,
                             cash_in_hand=cash_in_hand,
                             chart_labels=chart_labels,
                             chart_income=chart_income,
                             chart_expense=chart_expense,
                             installments=Installment.query.order_by(Installment.created_at).all(),
                             total_party_due=total_party_due,
                             total_party_advance=total_party_advance,
                             today_todos=today_todos,
                             sync_mismatch=current_app.config.get('SYNC_MISMATCH'))
    except Exception as e:
        print(f"CRITICAL ERROR in index view: {e}")
        return render_template('error.html', error=str(e), path='/'), 500

def verify_password():
    password = request.form.get('admin_password')
    if current_user.is_authenticated and current_user.check_password(password):
        return True
    # Fallback to old config password for now (transition period)
    admin_pass = current_app.config.get('ADMIN_PASSWORD')
    if password == admin_pass:
        return True
    return False

@main.before_app_request
def global_request_processor():
    # 1. Skip for all static assets and background sync requests
    if not request.endpoint or \
       request.endpoint == 'static' or \
       request.path.endswith(('.ico', '.png', '.jpg', '.json', '.js', '.css')) or \
       request.path.startswith('/static/'):
        return

    # 2. Define Public routes (accessible without login)
    public_endpoints = ['main.login', 'main.forgot_password', 'main.reset_password', 'main.uploaded_file']
    
    # Allow access to public routes
    if request.endpoint in public_endpoints:
        return

    # 3. Enforce Authentication
    if not current_user.is_authenticated:
        return redirect(url_for('main.login'))

    # 4. Sync mismatch check is now handled via a non-blocking banner on the dashboard
    # to avoid all possible redirect loops.
    return

@main.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('main.index'))
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        from models import User
        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            login_user(user)
            return redirect(url_for('main.index'))
        else:
            flash('Invalid username or password', 'danger')
    return render_template('login.html')

@main.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('main.login'))

@main.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        username = request.form.get('username')
        from models import User
        user = User.query.filter_by(username=username).first()
        if user:
            otp = ''.join(random.choices(string.digits, k=6))
            otp_store[username] = {
                'otp': otp,
                'expires': datetime.now().timestamp() + 600 # 10 mins
            }
            message = f"🔒 Password Reset OTP for {username}: {otp}\nExpires in 10 minutes."
            send_telegram_message(message)
            flash('Search your Telegram for the reset code.', 'info')
            return redirect(url_for('main.reset_password', username=username))
        else:
            flash('Username not found.', 'danger')
    return render_template('forgot_password.html')

@main.route('/reset-password', methods=['GET', 'POST'])
def reset_password():
    username = request.args.get('username') or request.form.get('username')
    if request.method == 'POST':
        otp_input = request.form.get('otp')
        new_password = request.form.get('new_password')
        
        entry = otp_store.get(username)
        if entry and entry['otp'] == otp_input and entry['expires'] > datetime.now().timestamp():
            from models import User
            user = User.query.filter_by(username=username).first()
            if user:
                user.set_password(new_password)
                db.session.commit()
                del otp_store[username]
                flash('Password reset successful! Please login.', 'success')
                return redirect(url_for('main.login'))
        else:
            flash('Invalid or expired OTP.', 'danger')
            
    return render_template('reset_password.html', username=username)

@main.route('/todos', methods=['GET', 'POST'])
@login_required
def manage_todos():
    from models import Todo
    if request.method == 'POST':
        task_text = request.form.get('task')
        category = request.form.get('category', 'General')
        due_date = request.form.get('due_date')
        due_time = request.form.get('due_time')
        
        new_todo = Todo(
            task=task_text,
            category=category,
            due_date=due_date,
            due_time=due_time,
            user_id=current_user.id
        )
        db.session.add(new_todo)
        db.session.commit()
        flash('Task added successfully!', 'success')
        return redirect(url_for('main.manage_todos'))
    
    todos = Todo.query.filter_by(user_id=current_user.id).order_by(Todo.created_at.desc()).all()
    return render_template('todos.html', todos=todos)

@main.route('/todos/toggle/<int:id>', methods=['POST'])
@login_required
def toggle_todo(id):
    from models import Todo
    todo = Todo.query.get_or_404(id)
    if todo.user_id != current_user.id:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    todo.is_completed = not todo.is_completed
    db.session.commit()
    return jsonify({'success': True, 'new_status': todo.is_completed})

@main.route('/todos/delete/<int:id>', methods=['POST'])
@login_required
def delete_todo(id):
    from models import Todo
    todo = Todo.query.get_or_404(id)
    if todo.user_id != current_user.id:
        flash('Unauthorized deletion attempt.', 'danger')
        return redirect(url_for('main.manage_todos'))
    
    db.session.delete(todo)
    db.session.commit()
    flash('Task deleted.', 'info')
    return redirect(url_for('main.manage_todos'))

@main.route('/change_password_request', methods=['GET', 'POST'])
def change_password_request():
    if request.method == 'POST':
        # Generate OTP
        otp = ''.join(random.choices(string.digits, k=6))
        
        if send_telegram_message(f"Your OTP for Password Change is: {otp}"):
            otp_store['current_otp'] = otp
            flash('OTP sent to Telegram!', 'info')
            return redirect(url_for('main.verify_otp_page'))
        else:
            flash('Failed to send OTP. Check internet or bot config.', 'danger')
            
    return render_template('change_password_request.html')

@main.route('/verify_otp', methods=['GET', 'POST'])
def verify_otp_page():
    if request.method == 'POST':
        user_otp = request.form.get('otp')
        new_password = request.form.get('new_password')
        
        if 'current_otp' in otp_store and otp_store['current_otp'] == user_otp:
            # Update Password
            # Use DATA_FOLDER to ensure persistence
            data_folder = current_app.config.get('DATA_FOLDER', current_app.root_path)
            config_path = os.path.join(data_folder, 'admin_config.json')
            
            try:
                with open(config_path, 'w') as f:
                    json.dump({"ADMIN_PASSWORD": new_password}, f)
            except Exception as e:
                flash(f'Error saving password: {e}', 'danger')
                return redirect(url_for('main.index'))
            
            # Update Runtime Config
            current_app.config['ADMIN_PASSWORD'] = new_password
            
            # Clear OTP
            del otp_store['current_otp']
            
            flash('Password Changed Successfully!', 'success')
            return redirect(url_for('main.index'))
        else:
            flash('Invalid OTP!', 'danger')
            
    return render_template('verify_otp.html')

# --- Employee Routes ---
@main.route('/employees')
def manage_employees():
    employees = Employee.query.order_by(db.cast(Employee.employee_id, db.Integer)).all()
    return render_template('employees.html', employees=employees)

@main.route('/employee/add', methods=['GET', 'POST'])
def add_employee():
    if request.method == 'POST':
        if not verify_password():
            flash('Invalid Admin Password!', 'danger')
            return redirect(url_for('main.add_employee'))
        
        emp_id = request.form.get('employee_id')
        name = request.form.get('name')
        
        # Check if ID exists
        if Employee.query.filter_by(employee_id=emp_id).first():
            flash('Employee ID already exists!', 'danger')
            return redirect(url_for('main.add_employee'))
            
        new_emp = Employee(
            employee_id=emp_id,
            name=name,
            designation=request.form.get('designation'),
            department=request.form.get('department'),
            joining_date=request.form.get('joining_date'),
            phone=request.form.get('phone'),
            net_salary=float(request.form.get('net_salary') or 0),
            cl_total=int(request.form.get('cl_total') or 0),
            ml_total=int(request.form.get('ml_total') or 0),
            fl_total=int(request.form.get('fl_total') or 0),
            el_total=int(request.form.get('el_total') or 0)
        )
        db.session.add(new_emp)
        db.session.commit()
        trigger_excel_sync()
        trigger_sync()
        backup_to_telegram("Added Employee: " + name)
        flash('Employee added successfully', 'success')
        return redirect(url_for('main.manage_employees'))
        
    return render_template('employee_form.html')

@main.route('/employee/edit/<int:id>', methods=['GET', 'POST'])
def edit_employee(id):
    employee = Employee.query.get_or_404(id)
    if request.method == 'POST':
        if not verify_password():
            flash('Invalid Admin Password!', 'danger')
            return redirect(url_for('main.edit_employee', id=id))
            
        employee.name = request.form.get('name')
        employee.designation = request.form.get('designation')
        employee.department = request.form.get('department')
        employee.joining_date = request.form.get('joining_date')
        employee.phone = request.form.get('phone')
        employee.net_salary = float(request.form.get('net_salary') or 0)
        employee.cl_total = int(request.form.get('cl_total') or 0)
        employee.ml_total = int(request.form.get('ml_total') or 0)
        employee.fl_total = int(request.form.get('fl_total') or 0)
        employee.el_total = int(request.form.get('el_total') or 0)
        
        db.session.commit()
        trigger_excel_sync()
        trigger_sync()
        backup_to_telegram("Edited Employee: " + employee.name)
        flash('Employee updated successfully', 'success')
        return redirect(url_for('main.manage_employees'))
        
    return render_template('employee_form.html', employee=employee)

@main.route('/employee/delete/<int:id>', methods=['POST'])
def delete_employee(id):
    if not verify_password():
        flash('Invalid Admin Password!', 'danger')
        return redirect(url_for('main.manage_employees'))
        
    employee = Employee.query.get_or_404(id)
    employee_name = employee.name
    db.session.delete(employee)
    db.session.commit()
    trigger_excel_sync()
    trigger_sync()
    backup_to_telegram("Deleted Employee: " + employee_name)
    flash('Employee deleted successfully', 'info')
    return redirect(url_for('main.manage_employees'))

# --- Attendance Routes ---

def calculate_working_hours(in_time_str, out_time_str):
    """Calculates difference in hours between two HH:MM strings."""
    if not in_time_str or not out_time_str:
        return 0.0
    try:
        in_time = datetime.strptime(in_time_str, '%H:%M')
        out_time = datetime.strptime(out_time_str, '%H:%M')
        # Handle overnight shifts if out_time is before in_time
        if out_time < in_time:
            out_time = out_time + pd.Timedelta(days=1)
        diff = out_time - in_time
        return diff.total_seconds() / 3600.0
    except ValueError:
        return 0.0

@main.route('/attendance')
def view_attendance():
    selected_date = request.args.get('date', datetime.today().strftime('%Y-%m-%d'))
    selected_employee = request.args.get('employee_id')
    
    query = Attendance.query
    if selected_date:
        query = query.filter_by(date=selected_date)
    if selected_employee:
        query = query.filter_by(employee_id=selected_employee)
        
    records = query.order_by(Attendance.date.desc(), Attendance.id).all()
    employees = Employee.query.order_by(db.cast(Employee.employee_id, db.Integer)).all()
    
    return render_template('attendance.html', records=records, employees=employees, 
                           selected_date=selected_date, selected_employee=selected_employee)

@main.route('/attendance/mark', methods=['GET', 'POST'])
def mark_attendance():
    employees = Employee.query.order_by(db.cast(Employee.employee_id, db.Integer)).all()
    today = datetime.today().strftime('%Y-%m-%d')
    
    if request.method == 'POST':
        if not verify_password():
            flash('Invalid Admin Password!', 'danger')
            return redirect(url_for('main.mark_attendance'))
            
        date = request.form.get('date')
        if not date:
            flash('Date is required!', 'danger')
            return redirect(url_for('main.mark_attendance'))
            
        # Get list of employee IDs submitted
        emp_ids = request.form.getlist('employee_id[]')
        count = 0
        
        for emp_id in emp_ids:
            status = request.form.get(f'status_{emp_id}')
            in_time = request.form.get(f'in_time_{emp_id}')
            out_time = request.form.get(f'out_time_{emp_id}')
            notes = request.form.get(f'notes_{emp_id}')
            
            # Skip if status is empty (though it Shouldn't be)
            if not status:
                continue
                
            if status in ['Absent', 'Leave', 'Off Day', 'FL']:
                in_time = None
                out_time = None
                
            working_hours = calculate_working_hours(in_time, out_time)
            
            # Check if record already exists for this date and employee
            existing = Attendance.query.filter_by(employee_id=emp_id, date=date).first()
            if existing:
                existing.status = status
                existing.in_time = in_time
                existing.out_time = out_time
                existing.working_hours = working_hours
                existing.notes = notes
            else:
                new_att = Attendance(
                    employee_id=emp_id,
                    date=date,
                    status=status,
                    in_time=in_time,
                    out_time=out_time,
                    working_hours=working_hours,
                    notes=notes
                )
                db.session.add(new_att)
            count += 1
            
        db.session.commit()
        trigger_excel_sync()
        trigger_sync()
        backup_to_telegram(f"Marked Attendance for {date} ({count} employees)")
        flash(f'Attendance marked successfully for {count} employees on {date}.', 'success')
        return redirect(url_for('main.view_attendance', date=date))
        
    return render_template('mark_attendance.html', employees=employees, today=today)

@main.route('/report/individual', methods=['GET', 'POST'])
def individual_report():
    employees = Employee.query.order_by(db.cast(Employee.employee_id, db.Integer)).all()
    
    # Defaults
    today = datetime.today()
    selected_month_year = today.strftime('%Y-%m')
    selected_emp_id = None
    report_data = None
    stats = {}
    emp = None
    
    # Global Print Headers
    company_name = "Company Name"
    company_address = ""
    data_dir = current_app.config.get('DATA_FOLDER', '.')
    settings_path = os.path.join(data_dir, 'company_settings.json')
    if os.path.exists(settings_path):
        try:
            with open(settings_path, 'r') as f:
                settings_data = json.load(f)
                company_name = settings_data.get('company_name', company_name)
                company_address = settings_data.get('company_address', company_address)
        except Exception:
            pass
            
    if request.method == 'POST':
        selected_emp_id = request.form.get('employee_id')
        selected_month_year = request.form.get('month_year', selected_month_year)
        
        if selected_emp_id and selected_month_year:
            year, month = selected_month_year.split('-')
            emp = Employee.query.get(selected_emp_id)
            if emp:
                _, days_in_month = calendar.monthrange(int(year), int(month))
                start_date = f"{year}-{month}-01"
                end_date = f"{year}-{month}-{days_in_month:02d}"
                
                # Fetch attendance for this month
                attendances = Attendance.query.filter(
                    Attendance.employee_id == emp.id,
                    Attendance.date >= start_date,
                    Attendance.date <= end_date
                ).all()
                
                # Build a dictionary by date for quick lookup
                att_dict = {a.date: a for a in attendances}

                # Fetch approved/pending leaves for this employee in this month
                from models import Leave as LeaveModel
                emp_leaves = LeaveModel.query.filter(
                    LeaveModel.employee_id == emp.id,
                    LeaveModel.status != 'Rejected'
                ).all()

                # Build a dict: date_str -> leave_type for every day covered by a leave
                leave_dict = {}
                for lv in emp_leaves:
                    try:
                        lv_from = datetime.strptime(lv.from_date, '%Y-%m-%d')
                        lv_to = datetime.strptime(lv.to_date, '%Y-%m-%d')
                        cur = lv_from
                        while cur <= lv_to:
                            leave_dict[cur.strftime('%Y-%m-%d')] = lv.leave_type
                            from datetime import timedelta
                            cur += timedelta(days=1)
                    except Exception:
                        pass

                report_data = []
                # Initialize counters
                stats = {
                    'Present': 0, 'Late': 0, 'Absent': 0, 'Leave': 0, 'Off Day': 0, 'FL': 0, 'EL': 0, 'CL': 0, 'ML': 0
                }
                
                import calendar as pycal
                for day in range(1, days_in_month + 1):
                    current_date_obj = datetime(int(year), int(month), day)
                    date_str = current_date_obj.strftime('%Y-%m-%d')
                    day_name = current_date_obj.strftime('%A')
                    
                    att_record = att_dict.get(date_str)
                    
                    if att_record:
                        status = att_record.status
                        in_time = att_record.in_time or '-'
                        out_time = att_record.out_time or '-'
                        working_hours = att_record.working_hours or 0.0
                        notes = att_record.notes or ''
                        
                        # If status is generic 'Leave', check the Leave table for specifics (FL, EL, CL, ML)
                        if status == 'Leave':
                            if date_str in leave_dict:
                                status = leave_dict[date_str]
                            elif notes and ':' in notes:
                                # Fallback: Check if notes start with CL:, ML:, FL:, EL:
                                prefix = notes.split(':')[0].strip().upper()
                                if prefix in ['CL', 'ML', 'FL', 'EL']:
                                    status = prefix
                        
                        # Increment exact status
                        if status in stats:
                            stats[status] += 1
                        else:
                            # if it's ML, CL etc (which are usually grouped under Leave)
                            if 'Leave' not in stats: stats['Leave'] = 0
                            stats['Leave'] += 1
                            
                        report_data.append({
                            'date': date_str,
                            'day_name': day_name,
                            'status': status,
                            'in_time': in_time,
                            'out_time': out_time,
                            'working_hours': working_hours,
                            'notes': notes
                        })
                    elif date_str in leave_dict:
                        # Day is covered by an approved/pending leave
                        leave_type = leave_dict[date_str]
                        if leave_type == 'FL':
                            stats['FL'] += 1
                        elif leave_type == 'EL':
                            stats['EL'] += 1
                        elif leave_type == 'CL':
                            stats['CL'] += 1
                        elif leave_type == 'ML':
                            stats['ML'] += 1
                        else:
                            stats['Leave'] += 1
                        report_data.append({
                            'date': date_str,
                            'day_name': day_name,
                            'status': leave_type,
                            'in_time': '-',
                            'out_time': '-',
                            'working_hours': 0.0,
                            'notes': ''
                        })
                    else:
                        # Auto-detect off days if not marked and is Friday
                        is_friday = (current_date_obj.weekday() == pycal.FRIDAY)
                        if is_friday:
                            status = 'Off Day'
                            stats['Off Day'] += 1
                        else:
                            # If no record and past today, it's Absent? We won't penalize future dates.
                            # Usually if no record, we just show empty or '-'
                            status = '-'
                            if current_date_obj <= today:
                                status = 'Unmarked'
                                
                        report_data.append({
                            'date': date_str,
                            'day_name': day_name,
                            'status': status,
                            'in_time': '-',
                            'out_time': '-',
                            'working_hours': 0.0,
                            'notes': ''
                        })


    display_month = ""
    if selected_month_year:
        try:
            y, m = selected_month_year.split('-')
            display_month = f"{calendar.month_name[int(m)]} {y}"
        except:
            pass

    return render_template('individual_report.html', 
                           employees=employees,
                           selected_month_year=selected_month_year,
                           selected_emp_id=selected_emp_id,
                           report_data=report_data,
                           stats=stats,
                           emp=emp,
                           display_month=display_month,
                           company_name=company_name,
                           company_address=company_address)

@main.route('/report/individual/export/excel', methods=['POST'])
def individual_report_export_excel():
    """Export individual attendance report to Excel."""
    import io
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from flask import send_file

    selected_emp_id = request.form.get('employee_id')
    selected_month_year = request.form.get('month_year')

    if not selected_emp_id or not selected_month_year:
        flash('Please select employee and month first.', 'warning')
        return redirect(url_for('main.individual_report'))

    year, month = selected_month_year.split('-')
    emp = Employee.query.get(selected_emp_id)
    if not emp:
        flash('Employee not found.', 'danger')
        return redirect(url_for('main.individual_report'))

    _, days_in_month = calendar.monthrange(int(year), int(month))
    start_date = f"{year}-{month}-01"
    end_date = f"{year}-{month}-{days_in_month:02d}"
    today = datetime.today()

    # Fetch attendance for this month
    from models import Attendance
    attendances = Attendance.query.filter(
        Attendance.employee_id == emp.id,
        Attendance.date >= start_date,
        Attendance.date <= end_date
    ).all()
    
    # Build a dictionary by date for quick lookup
    att_dict = {a.date: a for a in attendances}

    # Fetch approved/pending leaves for this employee in this month
    from models import Leave as LeaveModel
    emp_leaves = LeaveModel.query.filter(
        LeaveModel.employee_id == emp.id,
        LeaveModel.status != 'Rejected'
    ).all()

    # Build a dict: date_str -> leave_type for every day covered by a leave
    leave_dict = {}
    for lv in emp_leaves:
        try:
            lv_from = datetime.strptime(lv.from_date, '%Y-%m-%d')
            lv_to = datetime.strptime(lv.to_date, '%Y-%m-%d')
            cur = lv_from
            while cur <= lv_to:
                leave_dict[cur.strftime('%Y-%m-%d')] = lv.leave_type
                from datetime import timedelta
                cur += timedelta(days=1)
        except Exception:
            pass

    stats = {'Present': 0, 'Late': 0, 'Absent': 0, 'Leave': 0, 'Off Day': 0, 'FL': 0, 'EL': 0, 'CL': 0, 'ML': 0}
    report_data = []
    import calendar as pycal

    for day in range(1, days_in_month + 1):
        current_date_obj = datetime(int(year), int(month), day)
        date_str = current_date_obj.strftime('%Y-%m-%d')
        day_name = current_date_obj.strftime('%A')
        att_record = att_dict.get(date_str)

        if att_record:
            status = att_record.status
            in_time = att_record.in_time or '-'
            out_time = att_record.out_time or '-'
            working_hours = att_record.working_hours or 0.0
            
            # If status is generic 'Leave', check the Leave table for specifics (FL, EL, CL, ML)
            if status == 'Leave':
                if date_str in leave_dict:
                    status = leave_dict[date_str]
                elif att_record.notes and ':' in att_record.notes:
                    # Fallback: Check if notes start with CL:, ML:, FL:, EL:
                    prefix = att_record.notes.split(':')[0].strip().upper()
                    if prefix in ['CL', 'ML', 'FL', 'EL']:
                        status = prefix
            
            if status in stats:
                stats[status] += 1
            else:
                stats['Leave'] += 1
        elif date_str in leave_dict:
            # No attendance record, but a leave record exists
            status = leave_dict[date_str]
            if status in stats:
                stats[status] += 1
            else:
                stats['Leave'] += 1
            in_time = '-'
            out_time = '-'
            working_hours = 0.0
        else:
            is_friday = (current_date_obj.weekday() == pycal.FRIDAY)
            if is_friday:
                status = 'Off Day'
                stats['Off Day'] += 1
            elif current_date_obj <= today:
                status = 'Unmarked'
            else:
                status = '-'
            in_time = '-'
            out_time = '-'
            working_hours = 0.0

        report_data.append({
            'date': date_str,
            'day_name': day_name,
            'status': status,
            'in_time': in_time,
            'out_time': out_time,
            'working_hours': working_hours
        })

    # Load company info
    company_name = 'NEXUS RIVER VIEW'
    company_address = ''
    data_dir = current_app.config.get('DATA_FOLDER', '.')
    settings_path = os.path.join(data_dir, 'company_settings.json')
    if os.path.exists(settings_path):
        try:
            with open(settings_path, 'r') as f:
                s = json.load(f)
                company_name = s.get('company_name', company_name)
                company_address = s.get('company_address', company_address)
        except Exception:
            pass

    display_month = calendar.month_name[int(month)] + ' ' + year

    # Build Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Attendance Report'

    header_font = Font(bold=True, size=14)
    sub_header_font = Font(bold=True, size=11)
    bold_font = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center')
    header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
    header_font_white = Font(bold=True, color='FFFFFF')

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Company header
    ws.merge_cells('A1:F1')
    ws['A1'] = company_name
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = center

    ws.merge_cells('A2:F2')
    ws['A2'] = company_address
    ws['A2'].alignment = center

    ws.merge_cells('A3:F3')
    ws['A3'] = f'Monthly Attendance Report — {display_month}'
    ws['A3'].font = Font(bold=True, size=13)
    ws['A3'].alignment = center

    # Employee info
    ws['A5'] = 'Employee ID:'
    ws['B5'] = emp.employee_id
    ws['D5'] = 'Designation:'
    ws['E5'] = emp.designation or '-'
    ws['A6'] = 'Name:'
    ws['B6'] = emp.name
    ws['D6'] = 'Department:'
    ws['E6'] = emp.department or '-'

    for row in [5, 6]:
        for col in ['A', 'D']:
            ws[f'{col}{row}'].font = bold_font

    # Summary row
    summary_row = 8
    ws[f'A{summary_row}'] = 'Present'
    ws[f'B{summary_row}'] = stats['Present']
    ws[f'C{summary_row}'] = 'Absent'
    ws[f'D{summary_row}'] = stats['Absent']
    ws[f'E{summary_row}'] = 'Late'
    ws[f'F{summary_row}'] = stats['Late']

    summary_row2 = 9
    ws[f'A{summary_row2}'] = 'Off Day (OD)'
    ws[f'B{summary_row2}'] = stats['Off Day']
    ws[f'C{summary_row2}'] = 'Leave'
    ws[f'D{summary_row2}'] = stats['Leave']
    ws[f'E{summary_row2}'] = 'Festival (FL)'
    ws[f'F{summary_row2}'] = stats['FL']
    
    summary_row3 = 10
    ws[f'A{summary_row3}'] = 'Earned (EL)'
    ws[f'B{summary_row3}'] = stats['EL']
    ws[f'C{summary_row3}'] = 'Casual (CL)'
    ws[f'D{summary_row3}'] = stats['CL']
    ws[f'E{summary_row3}'] = 'Medical (ML)'
    ws[f'F{summary_row3}'] = stats['ML']

    for row in [summary_row, summary_row2, summary_row3]:
        for col in ['A', 'C', 'E']:
            if ws[f'{col}{row}'].value:
                ws[f'{col}{row}'].font = bold_font

    # Table headers
    table_start = 12
    headers = ['Date', 'Day', 'In Time', 'Out Time', 'Status', 'Work Hrs']
    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=table_start, column=idx, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # Data rows
    STATUS_COLORS = {
        'Present': 'D1FAE5',
        'Absent': 'FEE2E2',
        'Late': 'FEF3C7',
        'Off Day': 'E0F2FE',
        'Leave': 'F3F4F6',
        'FL': 'DBEAFE',
        'EL': 'DBEAFE',
    }

    for r_idx, row in enumerate(report_data, start=table_start + 1):
        vals = [row['date'], row['day_name'], row['in_time'], row['out_time'],
                row['status'], row['working_hours'] if row['working_hours'] else '-']
        for c_idx, val in enumerate(vals, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = center
            cell.border = thin_border
            status = row['status']
            if status in STATUS_COLORS:
                cell.fill = PatternFill(start_color=STATUS_COLORS[status],
                                        end_color=STATUS_COLORS[status], fill_type='solid')

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 13
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 12

    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Sanitize company name for filename
    safe_company_name = "".join(c for c in company_name if c.isalnum() or c in (' ', '_', '-')).strip().replace(' ', '_')
    filename = f'{safe_company_name}_Attendance_{emp.employee_id}_{emp.name}_{display_month}.xlsx'.replace(' ', '_')
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@main.route('/attendance/edit/<int:id>', methods=['POST'])
def edit_attendance(id):
    if not verify_password():
        flash('Invalid Admin Password!', 'danger')
        return redirect(url_for('main.view_attendance'))
        
    record = Attendance.query.get_or_404(id)
    record.status = request.form.get('status')
    
    if record.status in ['Absent', 'Leave', 'Off Day', 'FL']:
        record.in_time = None
        record.out_time = None
        record.working_hours = 0.0
    else:
        record.in_time = request.form.get('in_time')
        record.out_time = request.form.get('out_time')
        record.working_hours = calculate_working_hours(record.in_time, record.out_time)
        
    record.notes = request.form.get('notes')
    
    db.session.commit()
    trigger_excel_sync()
    trigger_sync()
    backup_to_telegram(f"Edited Attendance for {record.employee.name} on {record.date}")
    flash('Attendance updated successfully', 'success')
    return redirect(url_for('main.view_attendance', date=record.date))

@main.route('/attendance/delete/<int:id>', methods=['POST'])
def delete_attendance(id):
    if not verify_password():
        flash('Invalid Admin Password!', 'danger')
        return redirect(url_for('main.view_attendance'))
        
    record = Attendance.query.get_or_404(id)
    date_val = record.date
    name_val = record.employee.name
    
    db.session.delete(record)
    db.session.commit()
    trigger_excel_sync()
    trigger_sync()
    backup_to_telegram(f"Deleted Attendance for {name_val} on {date_val}")
    flash('Attendance record deleted.', 'info')
    return redirect(url_for('main.view_attendance', date=date_val))

# --- Leave Management Routes ---

@main.route('/leaves')
def view_leaves():
    leaves = Leave.query.order_by(Leave.created_at.desc()).all()
    employees = Employee.query.all()
    
    # Calculate Leave Balances
    leave_stats = []
    for emp in employees:
        emp_leaves = Leave.query.filter_by(employee_id=emp.id, status='Approved').all()
        cl_used = sum(l.total_days for l in emp_leaves if l.leave_type == 'CL')
        ml_used = sum(l.total_days for l in emp_leaves if l.leave_type == 'ML')
        fl_used = sum(l.total_days for l in emp_leaves if l.leave_type == 'FL')
        el_used = sum(l.total_days for l in emp_leaves if l.leave_type == 'EL')
        
        stat = {
            'employee': emp,
            'cl_total': emp.cl_total, 'cl_used': cl_used, 'cl_balance': emp.cl_total - cl_used,
            'ml_total': emp.ml_total, 'ml_used': ml_used, 'ml_balance': emp.ml_total - ml_used,
            'fl_total': emp.fl_total, 'fl_used': fl_used, 'fl_balance': emp.fl_total - fl_used,
            'el_total': emp.el_total, 'el_used': el_used, 'el_balance': emp.el_total - el_used,
        }
        leave_stats.append(stat)
        
    return render_template('leaves.html', leaves=leaves, leave_stats=leave_stats)

@main.route('/leave/apply', methods=['GET', 'POST'])
def apply_leave():
    if request.method == 'POST':
        if not verify_password():
            flash('Invalid Admin Password!', 'danger')
            return redirect(url_for('main.apply_leave'))
            
        from datetime import datetime, timedelta
        
        emp_ids = request.form.getlist('employee_ids')
        record_type = request.form.get('leave_type') # This is the "Record Type" dropdown
        from_date_str = request.form.get('from_date')
        to_date_str = request.form.get('to_date')
        reason = request.form.get('reason')
        status = request.form.get('status', 'Approved')
        
        # New Attendance-specific fields
        in_time = request.form.get('in_time', '11:00')
        out_time = request.form.get('out_time', '19:00')

        if not emp_ids:
            flash('Please select at least one employee.', 'warning')
            return redirect(url_for('main.apply_leave'))
            
        start_date = datetime.strptime(from_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(to_date_str, '%Y-%m-%d')
        
        days_processed = 0
        for emp_id in emp_ids:
            current_date = start_date
            while current_date <= end_date:
                curr_date_str = current_date.strftime('%Y-%m-%d')
                
                # OVERWRITE LOGIC: Delete pre-existing Attendance & Leave for this day
                # 1. Delete Attendance
                existing_att = Attendance.query.filter_by(employee_id=emp_id, date=curr_date_str).all()
                for att in existing_att:
                    db.session.delete(att)
                
                # 2. Delete overlapping Leave records for this specific DATE
                # Note: This is simpler if we just treat this new tool as the 'Master Override'
                # For simplicity, we search for any leave record that covers this date
                overlapping_leaves = Leave.query.filter(
                    Leave.employee_id == emp_id,
                    Leave.from_date <= curr_date_str,
                    Leave.to_date >= curr_date_str
                ).all()
                for l in overlapping_leaves:
                    # If it's a multi-day leave, we might technically need to split it, 
                    # but usually, users want a clean wipe if they are re-marking things.
                    db.session.delete(l)

                # CREATE NEW RECORD based on type
                if record_type in ['Present', 'Absent', 'Off Day']:
                    # Calculate working hours if Present
                    working_hours = 0.0
                    if record_type == 'Present':
                        try:
                            t1 = datetime.strptime(in_time, '%H:%M')
                            t2 = datetime.strptime(out_time, '%H:%M')
                            working_hours = (t2 - t1).seconds / 3600
                        except:
                            working_hours = 8.0 # Default fallback
                    
                    new_att = Attendance(
                        employee_id=emp_id,
                        date=curr_date_str,
                        in_time=in_time if record_type == 'Present' else '',
                        out_time=out_time if record_type == 'Present' else '',
                        status=record_type,
                        working_hours=working_hours,
                        notes=reason
                    )
                    db.session.add(new_att)
                else:
                    # It's a proper Leave type (CL, ML, FL, EL)
                    # For multi-day leaves in this loop, we handle them carefully.
                    # Actually, if the loop is per-day, we only need to add 1 Leave record 
                    # for the ENTIRE range once per employee, OR we add one per day as Attendance
                    # for the system to see it in salary sheets.
                    
                    # SYSTEM DESIGN CHOICE:
                    # We store the 'Leave' record for balance tracking, AND we store daily 'Attendance' 
                    # marked as 'Leave' so the salary generator picks it up.
                    
                    # We only add the Leave record ONCE per employee for the range
                    if current_date == start_date:
                        total_days = (end_date - start_date).days + 1
                        new_leave = Leave(
                            employee_id=emp_id,
                            leave_type=record_type,
                            from_date=from_date_str,
                            to_date=to_date_str,
                            total_days=total_days,
                            reason=reason,
                            status=status
                        )
                        db.session.add(new_leave)

                    # Mark Attendance as 'Leave' for every day in the range
                    if status == 'Approved':
                        new_att_leave = Attendance(
                            employee_id=emp_id,
                            date=curr_date_str,
                            status='Leave',
                            notes=f"{record_type}: {reason}"
                        )
                        db.session.add(new_att_leave)
                
                current_date += timedelta(days=1)
                days_processed += 1
            
        db.session.commit()
        trigger_excel_sync()
        trigger_sync()
        backup_to_telegram(f"Attendance/Leave Updated: {record_type} for {len(emp_ids)} staff over dates {from_date_str} to {to_date_str}")
        flash('Records updated successfully. Old entries for these dates were overwritten.', 'success')
        return redirect(url_for('main.view_leaves'))
        
    employees = Employee.query.order_by(db.cast(Employee.employee_id, db.Integer)).all()
    return render_template('apply_leave.html', employees=employees)

@main.route('/leave/delete/<int:id>', methods=['POST'])
def delete_leave(id):
    if not verify_password():
        flash('Invalid Admin Password!', 'danger')
        return redirect(url_for('main.view_leaves'))
        
    leave = Leave.query.get_or_404(id)
    emp_name = leave.employee.name  # read before detaching from session
    db.session.delete(leave)
    db.session.commit()
    trigger_excel_sync()
    trigger_sync()
    backup_to_telegram(f"Deleted Leave Application for {emp_name}")
    flash('Leave application deleted.', 'info')
    return redirect(url_for('main.view_leaves'))

@main.route('/leave/edit/<int:id>', methods=['POST'])
def edit_leave(id):
    if not verify_password():
        flash('Invalid Admin Password!', 'danger')
        return redirect(url_for('main.view_leaves'))
        
    leave = Leave.query.get_or_404(id)
    leave.status = request.form.get('status')
    leave.reason = request.form.get('reason')
    
    db.session.commit()
    trigger_excel_sync()
    trigger_sync()
    backup_to_telegram(f"Edited Leave for {leave.employee.name}")
    flash('Leave record updated.', 'success')
    return redirect(url_for('main.view_leaves'))

# --- Salary Routes ---

@main.route('/salary/sheet', methods=['GET'])
def view_salary_sheet():
    # Show current month by default
    today = datetime.today()
    selected_month_year = today.strftime('%Y-%m')
    
    # Check if we have records for this month
    year, month = selected_month_year.split('-')
    salaries = Salary.query.filter_by(year=year, month=month).join(Employee).order_by(db.cast(Employee.employee_id, db.Integer)).all()
    employees = Employee.query.order_by(db.cast(Employee.employee_id, db.Integer)).all()
    
    display_month = calendar.month_name[int(month)] + " " + year
    
    # Load company settings for print
    company_name = "Company Name"
    company_address = ""
    data_dir = current_app.config.get('DATA_FOLDER', '.')
    settings_path = os.path.join(data_dir, 'company_settings.json')
    if os.path.exists(settings_path):
        try:
            with open(settings_path, 'r') as f:
                settings_data = json.load(f)
                company_name = settings_data.get('company_name', company_name)
                company_address = settings_data.get('company_address', company_address)
        except Exception:
            pass
            
    return render_template('salary_sheet.html', 
                           selected_month_year=selected_month_year, 
                           display_month=display_month,
                           salaries=salaries,
                           employees=employees,
                           company_name=company_name,
                           company_address=company_address)

@main.route('/salary/sheet/export/excel', methods=['POST'])
def export_salary_sheet_excel():
    """Export the monthly salary sheet to Excel."""
    import io
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from flask import send_file
    
    month_year = request.form.get('month_year')
    if not month_year:
        flash('No month selected for export.', 'warning')
        return redirect(url_for('main.view_salary_sheet'))
        
    year, month = month_year.split('-')
    salaries = Salary.query.filter_by(year=year, month=month).join(Employee).order_by(db.cast(Employee.employee_id, db.Integer)).all()
    
    if not salaries:
        flash('No salary records found for this month to export.', 'warning')
        return redirect(url_for('main.view_salary_sheet'))
        
    display_month = calendar.month_name[int(month)] + " " + year
    
    # Load company settings
    company_name = "Nexus River View"
    company_address = ""
    data_dir = current_app.config.get('DATA_FOLDER', '.')
    settings_path = os.path.join(data_dir, 'company_settings.json')
    if os.path.exists(settings_path):
        try:
            with open(settings_path, 'r') as f:
                settings_data = json.load(f)
                company_name = settings_data.get('company_name', company_name)
                company_address = settings_data.get('company_address', company_address)
        except Exception:
            pass

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Salary Sheet {month_year}"
    
    # Styling
    center = Alignment(horizontal='center', vertical='center')
    bold_font = Font(bold=True)
    title_font = Font(bold=True, size=16)
    subtitle_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Header Section
    ws.merge_cells('A1:M1')
    ws['A1'] = company_name
    ws['A1'].font = title_font
    ws['A1'].alignment = center
    
    ws.merge_cells('A2:M2')
    ws['A2'] = company_address
    ws['A2'].alignment = center
    
    ws.merge_cells('A3:M3')
    ws['A3'] = f"Monthly Salary Sheet - {display_month}"
    ws['A3'].font = subtitle_font
    ws['A3'].alignment = center
    
    # Table Header
    headers = [
        'ID', 'Employee Name', 'Net Salary', 'Work Days', 'Present', 
        'Leave', 'OD', 'Absent', 'Late', 'Deduction', 
        'Mob. Bill', 'Bonus', 'Final Payable'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
        
    # Data Rows
    current_row = 6
    for s in salaries:
        ws.cell(row=current_row, column=1, value=s.employee.employee_id).border = thin_border
        ws.cell(row=current_row, column=2, value=s.employee.name).border = thin_border
        ws.cell(row=current_row, column=3, value=s.net_salary).border = thin_border
        ws.cell(row=current_row, column=4, value=s.working_days).border = thin_border
        ws.cell(row=current_row, column=5, value=s.present_days).border = thin_border
        ws.cell(row=current_row, column=6, value=s.leave_days).border = thin_border
        ws.cell(row=current_row, column=7, value=s.off_days).border = thin_border
        ws.cell(row=current_row, column=8, value=s.absent_days).border = thin_border
        ws.cell(row=current_row, column=9, value=s.late_days).border = thin_border
        ws.cell(row=current_row, column=10, value=s.deduction).border = thin_border
        ws.cell(row=current_row, column=11, value=s.mobile_bill).border = thin_border
        ws.cell(row=current_row, column=12, value=s.bonus).border = thin_border
        
        final_cell = ws.cell(row=current_row, column=13, value=s.final_salary)
        final_cell.font = bold_font
        final_cell.border = thin_border
        
        # Format currency columns
        for col in [3, 10, 11, 12, 13]:
            ws.cell(row=current_row, column=col).number_format = '#,##0.00'
            
        current_row += 1
        
    # Total Row
    total_row = current_row
    ws.cell(row=total_row, column=1, value='Total').font = bold_font
    ws.merge_cells(f'A{total_row}:B{total_row}')
    ws.cell(row=total_row, column=1).alignment = center
    
    # Sums
    col_mapping = {
        'net_salary': 3,
        'deduction': 10,
        'mobile_bill': 11,
        'bonus': 12,
        'final_salary': 13
    }
    
    for attr, col in col_mapping.items():
        total_val = sum(getattr(s, attr) or 0 for s in salaries)
        cell = ws.cell(row=total_row, column=col, value=total_val)
        cell.font = bold_font
        cell.number_format = '#,##0.00'
        cell.border = thin_border
    
    # Fill in borders for unused cells in total row
    for col in range(1, 14):
        if not ws.cell(row=total_row, column=col).border:
            ws.cell(row=total_row, column=col).border = thin_border

    # Column Widths
    ws.column_dimensions['B'].width = 25
    for col in ['C', 'J', 'K', 'L', 'M']:
        ws.column_dimensions[col].width = 15
        
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Sanitize company name for filename
    safe_company_name = "".join(c for c in company_name if c.isalnum() or c in (' ', '_', '-')).strip().replace(' ', '_')
    filename = f"{safe_company_name}_Salary_Sheet_{month_year}.xlsx"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@main.route('/salary/sheet/generate', methods=['GET', 'POST'])
def generate_salary_sheet():
    if request.method == 'GET':
        return redirect(url_for('main.view_salary_sheet'))
        
    action = request.form.get('action') # 'view' or 'generate'
    month_year = request.form.get('month_year') # format 'YYYY-MM'
    
    if not month_year:
        return redirect(url_for('main.view_salary_sheet'))
        
    year, monthstr = month_year.split('-')
    month_int = int(monthstr)
    
    # Load company settings for print
    company_name = "Company Name"
    company_address = "Company Address"
    data_dir = current_app.config.get('DATA_FOLDER', '.')
    settings_path = os.path.join(data_dir, 'company_settings.json')
    if os.path.exists(settings_path):
        try:
            with open(settings_path, 'r') as f:
                settings_data = json.load(f)
                company_name = settings_data.get('company_name', company_name)
                company_address = settings_data.get('company_address', company_address)
        except Exception:
            pass

    if action == 'view':
        salaries = Salary.query.filter_by(year=year, month=monthstr).join(Employee).order_by(db.cast(Employee.employee_id, db.Integer)).all()
        employees = Employee.query.order_by(db.cast(Employee.employee_id, db.Integer)).all()
        display_month = calendar.month_name[month_int] + " " + year
        return render_template('salary_sheet.html', 
                               selected_month_year=month_year, 
                               display_month=display_month,
                               salaries=salaries,
                               employees=employees,
                               company_name=company_name,
                               company_address=company_address)
                               
    # If action is generate
    if not verify_password():
        flash('Invalid Admin Password!', 'danger')
        return redirect(url_for('main.view_salary_sheet'))
        
    # Get total days in month
    _, total_days_in_month = calendar.monthrange(int(year), month_int)
    
    selected_emp_ids = request.form.getlist('employee_ids')
    if not selected_emp_ids:
        # Fallback to all employees if none selected for some reason
        employees_to_process = Employee.query.all()
    else:
        employees_to_process = Employee.query.filter(Employee.id.in_(selected_emp_ids)).all()
        
    count = 0
    
    for emp in employees_to_process:
        # Get start and end date strings for queries
        start_date = f"{year}-{monthstr}-01"
        end_date = f"{year}-{monthstr}-{total_days_in_month:02d}"
        
        # Pull attendance
        attendances = Attendance.query.filter(
            Attendance.employee_id == emp.id,
            Attendance.date >= start_date,
            Attendance.date <= end_date
        ).all()
        
        present_count = 0
        absent_count = 0
        leave_count = 0
        late_count = 0
        off_count = 0
        
        # Create a set of dates that have an explicit attendance record
        marked_dates = {datetime.strptime(att.date, '%Y-%m-%d').date() for att in attendances}
        
        # Analyze explicit attendance records for the month
        for att in attendances:
            if att.status == 'Present':
                present_count += 1
            elif att.status == 'Late':
                late_count += 1
            elif att.status == 'Absent':
                absent_count += 1
            elif att.status == 'Leave':
                leave_count += 1
            elif att.status == 'Off Day':
                off_count += 1
                
        # Automatically detect and count Fridays as Off Days IF they weren't explicitly marked otherwise
        import calendar as pycal
        for day in range(1, total_days_in_month + 1):
            current_date = datetime(int(year), month_int, day).date()
            if current_date.weekday() == pycal.FRIDAY and current_date not in marked_dates:
                off_count += 1
                
        # 3 Lates = 1 Absent Penalty
        late_penalty_absent = late_count // 3
        effective_absent = absent_count + late_penalty_absent
                
        # Unmarked days: Assume if a day is completely unmarked, it is not counted as absent or present yet?
        # Usually, companies mark absent for missing days. For safety, we only deduct for MARKED absent days.
        # This gives admins flexibility. 
        # Requirement: "If an employee is absent, the system should automatically deduct the per-day salary amount"
        # So we only deduct based on explicitly marked Absent days.
        
        deduction_per_day = emp.net_salary / total_days_in_month if total_days_in_month > 0 else 0
        total_deduction = effective_absent * deduction_per_day
        base_salary = emp.net_salary - total_deduction
        
        # Update or Create
        existing_salary = Salary.query.filter_by(employee_id=emp.id, month=monthstr, year=year).first()
        if existing_salary:
            # We don't overwrite if it's already Paid to prevent altering history, but we could.
            # Let's overwrite but keep status.
            existing_salary.net_salary = emp.net_salary
            existing_salary.working_days = total_days_in_month
            existing_salary.present_days = present_count + late_count # Late is counted as present physically
            existing_salary.absent_days = absent_count
            existing_salary.late_days = late_count
            existing_salary.leave_days = leave_count
            existing_salary.off_days = off_count
            existing_salary.per_day_salary = deduction_per_day
            existing_salary.deduction = total_deduction
            # calculate final with existing bonuses
            final_salary = max(0.0, base_salary + existing_salary.mobile_bill + existing_salary.bonus)
            existing_salary.final_salary = final_salary
        else:
            final_salary = max(0.0, base_salary)
            new_salary = Salary(
                employee_id=emp.id,
                month=monthstr,
                year=year,
                net_salary=emp.net_salary,
                working_days=total_days_in_month,
                present_days=present_count + late_count,
                absent_days=absent_count,
                leave_days=leave_count,
                late_days=late_count,
                off_days=off_count,
                per_day_salary=deduction_per_day,
                deduction=total_deduction,
                mobile_bill=0.0,
                bonus=0.0,
                final_salary=final_salary
            )
            db.session.add(new_salary)
        count += 1
        
    db.session.commit()
    trigger_excel_sync()
    trigger_sync()
    backup_to_telegram(f"Generated Salary Sheet for {month_year} ({count} employees)")
    flash(f'Salary Sheet generated for {count} employees.', 'success')
    
    # Render view after generation
    salaries = Salary.query.filter_by(year=year, month=monthstr).order_by(Salary.employee_id).all()
    display_month = calendar.month_name[month_int] + " " + year
    return render_template('salary_sheet.html', 
                           selected_month_year=month_year, 
                           display_month=display_month,
                           salaries=salaries,
                           company_name=company_name,
                           company_address=company_address)

@main.route('/salary/mark_paid/<int:id>', methods=['POST'])
def mark_salary_paid(id):
    # Optional security: if not verify_password(): ...
    # We will assume you perform this check if you want, else let admin just click it.
    salary = Salary.query.get_or_404(id)
    salary.status = 'Paid'
    salary.payment_date = datetime.today().strftime('%Y-%m-%d %H:%M:%S')
    db.session.commit()
    return redirect(url_for('main.generate_salary_sheet'))

@main.route('/salary/mark_unpaid/<int:id>', methods=['POST'])
def mark_salary_unpaid(id):
    if not verify_password():
        flash('Invalid Admin Password!', 'danger')
        return redirect(url_for('main.view_salary_sheet'))
        
    salary = Salary.query.get_or_404(id)
    salary.status = 'Unpaid'
    salary.payment_date = None
    db.session.commit()
    flash('Salary record reverted to Unpaid.', 'info')
    return redirect(url_for('main.view_salary_sheet'))

@main.route('/salary/edit/<int:id>', methods=['POST'])
def edit_salary(id):
    if not verify_password():
        flash('Invalid Admin Password!', 'danger')
        return redirect(url_for('main.generate_salary_sheet'))
        
    salary = Salary.query.get_or_404(id)
    try:
        salary.mobile_bill = float(request.form.get('mobile_bill', 0.0))
    except ValueError:
        pass
        
    try:
        salary.bonus = float(request.form.get('bonus', 0.0))
    except ValueError:
        pass
        
    # Recalculate final 
    base_salary = salary.net_salary - salary.deduction
    salary.final_salary = max(0.0, base_salary + salary.mobile_bill + salary.bonus)
    
    db.session.commit()
    trigger_excel_sync()
    trigger_sync()
    flash(f"Updated salary fields for {salary.employee.name}", 'success')
    # Because generate doesn't take args gracefully, we redirect back to view via POST
    # We can fake the POST by doing render_template instead, or redirecting to GET route
    return redirect(url_for('main.view_salary_sheet'))

@main.route('/salary/delete/<int:id>', methods=['POST'])
def delete_salary(id):
    if not verify_password():
        flash('Invalid Admin Password!', 'danger')
        return redirect(url_for('main.view_salary_sheet'))
        
    salary = Salary.query.get_or_404(id)
    name_val = salary.employee.name
    month_val = salary.month
    
    db.session.delete(salary)
    db.session.commit()
    trigger_excel_sync()
    trigger_sync()
    backup_to_telegram(f"Deleted Salary record for {name_val}")
    flash(f'Salary record for {name_val} deleted.', 'info')
    return redirect(url_for('main.view_salary_sheet'))

# --- Director Routes ---
@main.route('/director/add', methods=['GET', 'POST'])
def add_director():
    if request.method == 'POST':
        if not verify_password():
            flash('Invalid Admin Password!', 'danger')
            return redirect(url_for('main.add_director'))

        name = request.form.get('name')
        phone = request.form.get('phone')
        
        if name:
            new_director = Director(
                name=name, 
                phone=phone,
                bank_name=request.form.get('bank_name'),
                total_share=float(request.form.get('total_share') or 0),
                per_share_value=0,
                fair_cost=0,
                land_value_extra_share=0,
                total_paid=0,
                total_due=0,
                payment_history=request.form.get('payment_history')
            )
            # No calculation needed for new empty director
            db.session.add(new_director)
            db.session.commit()
            trigger_excel_sync()
            trigger_sync()
            backup_to_telegram("Added Director: " + name)
            flash('Director added successfully!', 'success')
            return redirect(url_for('main.index'))
    return render_template('director_form.html')

@main.route('/director/edit/<int:id>', methods=['GET', 'POST'])
def edit_director(id):
    director = Director.query.get_or_404(id)
    if request.method == 'POST':
        if not verify_password():
            flash('Invalid Admin Password!', 'danger')
            return redirect(url_for('main.edit_director', id=id))

        director.name = request.form.get('name')
        director.phone = request.form.get('phone')
        director.bank_name = request.form.get('bank_name')
        director.total_share = float(request.form.get('total_share') or 0)
        director.per_share_value = float(request.form.get('per_share_value') or 0)
        director.fair_cost = float(request.form.get('fair_cost') or 0)
        director.land_value_extra_share = float(request.form.get('land_value_extra_share') or 0)
        # total_paid and total_due are handled by background sync or manual recalculation
        
        # Recalculate Totals from Customers
        director.total_paid = sum(c.total_paid for c in director.customers)
        director.total_due = sum(c.due_amount for c in director.customers)

        db.session.commit()
        trigger_excel_sync()
        trigger_sync()
        backup_to_telegram("Edited Director: " + director.name)
        flash('Director updated successfully!', 'success')
        return redirect(url_for('main.index'))
    return render_template('director_form.html', director=director)

@main.route('/director/delete/<int:id>', methods=['POST'])
def delete_director(id):
    if not verify_password():
        flash('Invalid Admin Password!', 'danger')
        return redirect(url_for('main.index'))

    director = Director.query.get_or_404(id)
    # Optional: logic to handle orphaned customers (delete or unlink). 
    # For now, let's assume we delete them or db cascade (we didn't set cascade).
    # Ideally we should warn user. But per request, simple Remove.
    # Let's delete customers first manually if constraints exist
    for customer in director.customers:
        db.session.delete(customer)
        
    db.session.delete(director)
    db.session.commit()
    trigger_excel_sync()
    trigger_sync()
    backup_to_telegram("Deleted Director: " + director.name)
    flash('Director and their customers removed!', 'info')
    return redirect(url_for('main.index'))


# --- Customer Routes ---
@main.route('/customer/add', methods=['GET', 'POST'])
def add_customer():
    directors = Director.query.all()
    if request.method == 'POST':
        if not verify_password():
            flash('Invalid Admin Password!', 'danger')
            return redirect(url_for('main.add_customer'))

        # Extraction
        director_id = request.form.get('director_id')
        customer_id = request.form.get('customer_id')
        name = request.form.get('name')
        phone = request.form.get('phone')
        father_name = request.form.get('father_name')
        mother_name = request.form.get('mother_name')
        dob = request.form.get('dob')
        religion = request.form.get('religion')
        profession = request.form.get('profession')
        nid_no = request.form.get('nid_no')
        present_address = request.form.get('present_address')
        permanent_address = request.form.get('permanent_address')
        plot_no = request.form.get('plot_no')
        total_price = float(request.form.get('total_price') or 0)
        down_payment = float(request.form.get('down_payment') or 0)
        monthly_installment = float(request.form.get('monthly_installment') or 0)
        total_paid = float(request.form.get('total_paid') or 0)
        shares = float(request.form.get('shares') or 0)
        
        # Validation: Check Director Shares
        director = Director.query.get(director_id)
        if shares > director.available_shares:
            flash(f'Error: Director only has {director.available_shares} shares available.', 'danger')
            return redirect(url_for('main.add_customer'))

        if not customer_id:
            customer_id = generate_next_customer_id(director_id)

        # Calculation
        new_customer = Customer(
            director_id=director_id,
            customer_id=customer_id,
            name=name,
            phone=phone,
            father_name=father_name,
            mother_name=mother_name,
            dob=dob,
            religion=religion,
            profession=profession,
            nid_no=nid_no,
            present_address=present_address,
            permanent_address=permanent_address,
            plot_no=plot_no,
            total_price=total_price,
            down_payment=down_payment,
            monthly_installment=monthly_installment,
            total_paid=total_paid,
            due_amount=0, # Will be recalculated
            shares=shares
        )
        db.session.add(new_customer)
        db.session.flush() # Get ID
        
        # Initial recalculation
        recalculate_customer_totals(new_customer)
        
        db.session.commit()
        trigger_excel_sync()
        trigger_sync()
        backup_to_telegram("Added Customer: " + name)
        flash('Customer added successfully!', 'success')
        return redirect(url_for('main.index'))
        
    return render_template('customer_form.html', directors=directors)

@main.route('/customer/edit/<int:id>', methods=['GET', 'POST'])
def edit_customer(id):
    customer = Customer.query.get_or_404(id)
    directors = Director.query.all()
    
    if request.method == 'POST':
        if not verify_password():
            flash('Invalid Admin Password!', 'danger')
            return redirect(url_for('main.edit_customer', id=id))

        old_director = Director.query.get(customer.director_id)
        new_director_id = request.form.get('director_id')
        new_director = Director.query.get(new_director_id)
        new_shares = float(request.form.get('shares') or 0)

        # Validation: Check Available Shares (adjusting for current shares)
        available = new_director.available_shares + (customer.shares if old_director.id == new_director.id else 0)
        if new_shares > available:
            flash(f'Error: Director only has {available} shares available.', 'danger')
            return redirect(url_for('main.edit_customer', id=id))

        # Revert old values for Director
        old_director.total_paid -= customer.total_paid
        
        customer.director_id = new_director_id
        customer.customer_id = request.form.get('customer_id')
        customer.name = request.form.get('name')
        customer.phone = request.form.get('phone')
        customer.father_name = request.form.get('father_name')
        customer.mother_name = request.form.get('mother_name')
        customer.dob = request.form.get('dob')
        customer.religion = request.form.get('religion')
        customer.profession = request.form.get('profession')
        customer.nid_no = request.form.get('nid_no')
        customer.present_address = request.form.get('present_address')
        customer.permanent_address = request.form.get('permanent_address')
        customer.plot_no = request.form.get('plot_no')
        customer.total_price = float(request.form.get('total_price') or 0)
        customer.down_payment = float(request.form.get('down_payment') or 0)
        customer.monthly_installment = float(request.form.get('monthly_installment') or 0)
        customer.total_paid = float(request.form.get('total_paid') or 0)
        customer.shares = new_shares
        
        # Calc Due
        customer.due_amount = customer.total_price - customer.total_paid
        
        # Update New Director
        new_director.total_paid += customer.total_paid
        
        # Recalculate Both Directors
        for d in [old_director, new_director]:
            d.total_paid = sum(c.total_paid for c in d.customers)
            d.total_due = sum(c.due_amount for c in d.customers)

        db.session.commit()
        trigger_excel_sync()
        trigger_sync()
        backup_to_telegram("Edited Customer: " + customer.name)
        flash('Customer updated successfully!', 'success')
        return redirect(url_for('main.index'))
        
    return render_template('customer_form.html', customer=customer, directors=directors)

@main.route('/delete_customer/<int:id>', methods=['POST'])
def delete_customer(id):
    if not verify_password():
        flash('Invalid Admin Password!', 'danger')
        return redirect(url_for('main.index'))

    customer = Customer.query.get_or_404(id)
    director = Director.query.get(customer.director_id)
    
    # Update Director Totals
    director.total_paid = sum(c.total_paid for c in director.customers if c.id != customer.id)
    director.total_due = sum(c.due_amount for c in director.customers if c.id != customer.id)

    db.session.delete(customer)
    db.session.commit()
    # Sync to Excel
    trigger_excel_sync()
    trigger_sync()
    backup_to_telegram("Deleted Customer")
    flash('Customer deleted!', 'success')
    return redirect(url_for('main.index'))

# --- Installment Management ---
@main.route('/installment/create', methods=['POST'])
def create_installment():
    if not verify_password():
        flash('Invalid Admin Password!', 'danger')
        return redirect(url_for('main.index'))
    
    name = request.form.get('name')
    amount_per_share = float(request.form.get('amount_per_share') or 0)
    
    if name and amount_per_share > 0:
        new_inst = Installment(name=name, amount_per_share=amount_per_share)
        db.session.add(new_inst)
        db.session.flush() # Get ID
        
        # Generate for all customers
        customers = Customer.query.all()
        for c in customers:
            total_amt = c.shares * amount_per_share
            if total_amt > 0:
                cust_inst = CustomerInstallment(
                    customer_id=c.id,
                    installment_id=new_inst.id,
                    total_amount=total_amt,
                    due_amount=total_amt
                )
                db.session.add(cust_inst)
            # Ensure totals are correct for this customer
            recalculate_customer_totals(c)
        
        db.session.commit()
        flash(f'Installment "{name}" created for all applicable customers.', 'success')
    else:
        flash('Invalid Installment Name or Amount.', 'danger')
        
    return redirect(url_for('main.manage_installments_page'))

@main.route('/installment/edit/<int:id>', methods=['POST'])
def edit_installment(id):
    if not verify_password():
        flash('Invalid Admin Password!', 'danger')
        return redirect(url_for('main.index'))

    inst = Installment.query.get_or_404(id)
    new_name = request.form.get('name', '').strip()
    new_amount = float(request.form.get('amount_per_share') or 0)

    if not new_name or new_amount <= 0:
        flash('Invalid name or amount.', 'danger')
        return redirect(url_for('main.index'))

    inst.name = new_name
    inst.amount_per_share = new_amount

    # Recalculate all CustomerInstallment records for this installment
    for ci in inst.customer_installments:
        ci.total_amount = ci.customer.shares * new_amount
        ci.due_amount = ci.total_amount - ci.paid_amount

    db.session.flush()

    # Recalculate all affected customers
    affected_customers = set(ci.customer for ci in inst.customer_installments)
    for c in affected_customers:
        recalculate_customer_totals(c)

    # Recalculate all director dues using total_share
    total_rate_per_share = sum(i.amount_per_share for i in Installment.query.all())
    from models import Director as Dir
    for d in Dir.query.all():
        d.total_paid = sum(c.total_paid for c in d.customers)
        d.total_due = d.total_share * total_rate_per_share - d.total_paid

    db.session.commit()
    backup_to_telegram("Edited Installment")
    flash(f'Installment "{inst.name}" updated and all totals recalculated.', 'success')
    return redirect(url_for('main.manage_installments_page'))


@main.route('/installment/delete/<int:id>', methods=['POST'])
def delete_installment(id):
    if not verify_password():
        flash('Invalid Admin Password!', 'danger')
        return redirect(url_for('main.index'))

    inst = Installment.query.get_or_404(id)
    inst_name = inst.name

    # Collect affected customers before deleting
    affected_customers = set(ci.customer for ci in inst.customer_installments)

    # Delete the installment (cascade deletes CustomerInstallment records via model)
    db.session.delete(inst)
    db.session.flush()

    # Recalculate customer totals now that installment is removed
    for c in affected_customers:
        recalculate_customer_totals(c)

    # Recalculate all director dues
    total_rate_per_share = sum(i.amount_per_share for i in Installment.query.all())
    from models import Director as Dir
    for d in Dir.query.all():
        d.total_paid = sum(c.total_paid for c in d.customers)
        d.total_due = d.total_share * total_rate_per_share - d.total_paid

    db.session.commit()
    backup_to_telegram("Deleted Installment")
    flash(f'Installment "{inst_name}" deleted and all totals recalculated.', 'success')
    return redirect(url_for('main.manage_installments_page'))

@main.route('/installments')
def manage_installments_page():
    installments = Installment.query.order_by(Installment.created_at).all()
    return render_template('installments.html', installments=installments)

@main.route('/installment/<int:id>/details')
def installment_details(id):
    installment = Installment.query.get_or_404(id)
    # Sort by customer name for better readability
    customer_installments = sorted(installment.customer_installments, key=lambda x: x.customer.name)
    return render_template('installment_details.html', installment=installment, customer_installments=customer_installments)

# --- Party Category CRUD ---
@main.route('/party/category/add', methods=['POST'])
def add_party_category():
    name = request.form.get('name', '').strip()
    if name:
        existing = PartyCategory.query.filter_by(name=name).first()
        if not existing:
            db.session.add(PartyCategory(name=name))
            db.session.commit()
            flash(f'Category "{name}" added!', 'success')
        else:
            flash(f'Category "{name}" already exists.', 'warning')
    return redirect(url_for('main.manage_parties'))

@main.route('/party/category/edit/<int:id>', methods=['POST'])
def edit_party_category(id):
    cat = PartyCategory.query.get_or_404(id)
    new_name = request.form.get('name', '').strip()
    if new_name and new_name != cat.name:
        old_name = cat.name
        cat.name = new_name
        # Also update all existing Party entries with this category
        Party.query.filter_by(category=old_name).update({'category': new_name})
        db.session.commit()
        flash(f'Category renamed to "{new_name}".', 'success')
    return redirect(url_for('main.manage_parties'))

@main.route('/party/category/delete/<int:id>', methods=['POST'])
def delete_party_category(id):
    cat = PartyCategory.query.get_or_404(id)
    name = cat.name
    db.session.delete(cat)
    db.session.commit()
    flash(f'Category "{name}" deleted.', 'info')
    return redirect(url_for('main.manage_parties'))

# --- Party Ledger Routes ---
@main.route('/parties')
def manage_parties():
    # Seed defaults if empty
    if PartyCategory.query.count() == 0:
        defaults = ["Supplier", "Contractor", "Individual"]
        for d in defaults:
            db.session.add(PartyCategory(name=d))
        db.session.commit()

    parties = Party.query.order_by(Party.name).all()
    party_categories = PartyCategory.query.order_by(PartyCategory.name).all()
    return render_template('parties.html', parties=parties, party_categories=party_categories)

@main.route('/party/add', methods=['GET', 'POST'])
def add_party():
    if request.method == 'POST':
        if not verify_password():
            flash('Invalid Admin Password!', 'danger')
            return redirect(url_for('main.add_party'))
        
        name = request.form.get('name')
        category = request.form.get('category')
        phone = request.form.get('phone')
        address = request.form.get('address')
        
        if name and category:
            new_party = Party(name=name, category=category, phone=phone, address=address)
            db.session.add(new_party)
            db.session.commit()
            backup_to_telegram("Added Party: " + name)
            flash('Party added successfully!', 'success')
            return redirect(url_for('main.manage_parties'))
            
    categories = [c.name for c in PartyCategory.query.order_by(PartyCategory.name).all()]
    return render_template('party_form.html', categories=categories)

@main.route('/party/edit/<int:id>', methods=['GET', 'POST'])
def edit_party(id):
    party = Party.query.get_or_404(id)
    if request.method == 'POST':
        if not verify_password():
            flash('Invalid Admin Password!', 'danger')
            return redirect(url_for('main.edit_party', id=id))
            
        party.name = request.form.get('name')
        party.category = request.form.get('category')
        party.phone = request.form.get('phone')
        party.address = request.form.get('address')
        
        db.session.commit()
        backup_to_telegram("Edited Party: " + party.name)
        flash('Party updated successfully!', 'success')
        return redirect(url_for('main.manage_parties'))
        
    categories = [c.name for c in PartyCategory.query.order_by(PartyCategory.name).all()]
    return render_template('party_form.html', party=party, categories=categories)

@main.route('/party/delete/<int:id>', methods=['POST'])
def delete_party(id):
    if not verify_password():
        flash('Invalid Admin Password!', 'danger')
        return redirect(url_for('main.manage_parties'))
        
    party = Party.query.get_or_404(id)
    party_name = party.name
    db.session.delete(party)
    db.session.commit()
    backup_to_telegram("Deleted Party: " + party_name)
    flash('Party deleted successfully!', 'info')
    return redirect(url_for('main.manage_parties'))

@main.route('/party/ledger/<int:id>')
def view_party_ledger(id):
    party = Party.query.get_or_404(id)
    # Strict ordering to ensure the table matches the running balance calculation
    entries = PartyLedger.query.filter_by(party_id=id).order_by(PartyLedger.date.asc(), PartyLedger.id.asc()).all()
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    if start_date and end_date:
        entries = [e for e in entries if start_date <= e.date <= end_date]
        
    return render_template('party_ledger.html', party=party, entries=entries, start_date=start_date, end_date=end_date)

@main.route('/party/ledger/recalculate/<int:id>')
def recalculate_ledger(id):
    from logic import recalculate_party_ledger_balances
    if recalculate_party_ledger_balances(id):
        flash('Ledger balances recalculated successfully!', 'success')
    else:
        flash('Failed to recalculate ledger balances.', 'danger')
    return redirect(url_for('main.view_party_ledger', id=id))

@main.route('/party/ledger/add/<int:party_id>', methods=['POST'])
def add_ledger_entry(party_id):
    from routes import verify_password
    if not verify_password():
        flash('Invalid Admin Password!', 'danger')
        return redirect(url_for('main.view_party_ledger', id=party_id))
        
    date = request.form.get('date')
    description = request.form.get('description')
    bill_amount = float(request.form.get('bill_amount') or 0)
    paid_amount = float(request.form.get('paid_amount') or 0)
    reference = request.form.get('reference')
    
    new_entry = PartyLedger(
        party_id=party_id,
        date=date,
        description=description,
        bill_amount=bill_amount,
        paid_amount=paid_amount,
        balance=0.0, # Will be set by recalculate
        reference=reference
    )
    db.session.add(new_entry)
    db.session.commit()
    
    # Recalculate to ensure all balances are correct (including future dates if any)
    recalculate_party_ledger_balances(party_id)
    
    backup_to_telegram("Added Party Ledger Entry: " + description)
    
    flash('Ledger entry added!', 'success')
    return redirect(url_for('main.view_party_ledger', id=party_id))

@main.route('/party/ledger/edit/<int:id>', methods=['POST'])
def edit_ledger_entry(id):
    from routes import verify_password
    entry = PartyLedger.query.get_or_404(id)
    party_id = entry.party_id
    
    if not verify_password():
        flash('Invalid Admin Password!', 'danger')
        return redirect(url_for('main.view_party_ledger', id=party_id))
        
    entry.date = request.form.get('date')
    entry.description = request.form.get('description')
    entry.bill_amount = float(request.form.get('bill_amount') or 0)
    entry.paid_amount = float(request.form.get('paid_amount') or 0)
    entry.reference = request.form.get('reference')
    
    db.session.commit()
    
    # Recalculate balances after edit
    recalculate_party_ledger_balances(party_id)
    
    backup_to_telegram("Edited Party Ledger Entry")
    
    flash('Ledger entry updated!', 'success')
    return redirect(url_for('main.view_party_ledger', id=party_id))

@main.route('/party/ledger/delete/<int:id>', methods=['POST'])
def delete_ledger_entry(id):
    from routes import verify_password
    entry = PartyLedger.query.get_or_404(id)
    party_id = entry.party_id
    
    if not verify_password():
        flash('Invalid Admin Password!', 'danger')
        return redirect(url_for('main.view_party_ledger', id=party_id))
        
    db.session.delete(entry)
    db.session.commit()
    
    # Recalculate balances after deletion
    recalculate_party_ledger_balances(party_id)
    
    backup_to_telegram("Deleted Party Ledger Entry")
    
    flash('Ledger entry deleted!', 'success')
    return redirect(url_for('main.view_party_ledger', id=party_id))

@main.route('/party/ledger/export/<int:id>')
def export_party_ledger(id):
    output, filename = export_party_ledger_to_excel(id)
    if not output:
        flash(filename, 'danger')
        return redirect(url_for('main.manage_parties'))
        
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@main.route('/manage_transactions/<int:customer_id>', methods=['GET', 'POST'])
def manage_transactions(customer_id):
    customer = Customer.query.get_or_404(customer_id)
    
    if request.method == 'POST':
        if not verify_password():
            flash('Invalid Admin Password!', 'danger')
            return redirect(url_for('main.manage_transactions', customer_id=customer_id))

        date = request.form.get('date')
        amount = float(request.form.get('amount') or 0)
        installment_type = request.form.get('installment_type')
        cust_inst_id = request.form.get('customer_installment_id') or None
        transaction_id = request.form.get('transaction_id')
        remarks = request.form.get('remarks')
        payment_source = request.form.get('payment_source', 'cash')  # 'bank', 'petty_cash', or 'cash'
        
        # Payment source specific fields
        bank_id = request.form.get('bank_id') or None
        bank_name = request.form.get('bank_name', '')
        cheque_no = request.form.get('cheque_no', '')
        bank_narration = request.form.get('bank_narration', '')
        bank_remarks = request.form.get('bank_remarks', '')
        petty_cash_desc = request.form.get('petty_cash_desc', f'Payment from {customer.name}')
        petty_cash_category = request.form.get('petty_cash_category', 'Customer Payment')
        
        # Merge remarks: bank_remarks takes priority when bank is selected
        if payment_source == 'bank' and bank_remarks:
            remarks = bank_remarks
        
        # Handle Images
        images = []
        if 'evidence' in request.files:
            files = request.files.getlist('evidence')
            for file in files:
                    from image_utils import save_as_webp
                    filename = save_as_webp(file, current_app.config['UPLOAD_FOLDER'])
                    if filename:
                        images.append(filename)
        
        # Determine bank_name to store on transaction
        if payment_source == 'bank' and bank_id:
            selected_bank = Bank.query.get(bank_id)
            if selected_bank:
                bank_name = selected_bank.bank_name
        elif payment_source == 'petty_cash':
            bank_name = 'Petty Cash'
        
        new_tx = Transaction(
            date=date,
            amount=amount,
            installment_type=installment_type,
            bank_name=bank_name,
            transaction_id=transaction_id,
            remarks=remarks,
            customer_id=customer_id,
            customer_installment_id=cust_inst_id,
            images=','.join(images)
        )
        db.session.add(new_tx)
        
        # Update Customer Totals
        customer.total_paid += amount
        customer.due_amount = customer.total_price - customer.total_paid
        
        # Update Director Totals
        director = Director.query.get(customer.director_id)
        director.total_paid = sum(c.total_paid for c in director.customers)
        director.total_due = sum(c.due_amount for c in director.customers)

        # Update Specific Installment if selected
        if cust_inst_id:
            cust_inst = CustomerInstallment.query.get(cust_inst_id)
            if cust_inst:
                cust_inst.paid_amount += amount
                cust_inst.due_amount = cust_inst.total_amount - cust_inst.paid_amount
                # Mirror installment name if not set
                if not new_tx.installment_type:
                    new_tx.installment_type = cust_inst.installment.name

        # Handle Payment Source: update bank balance or petty cash
        if payment_source == 'bank' and bank_id:
            selected_bank = Bank.query.get(bank_id)
            if selected_bank:
                # Calculate running balance
                last_tx = BankTransaction.query.filter_by(bank_id=selected_bank.id).order_by(BankTransaction.id.desc()).first()
                running_balance = (last_tx.balance if last_tx else 0) + amount
                bank_tx = BankTransaction(
                    date=date,
                    cheque_no=cheque_no or None,
                    ref_no=transaction_id or None,
                    narration=bank_narration or f'Customer Payment - {customer.name} ({customer.customer_id})',
                    transaction_details=installment_type or 'Customer Payment',
                    credit=amount,
                    debit=0,
                    balance=running_balance,
                    bank_id=selected_bank.id
                )
                db.session.add(bank_tx)
        elif payment_source == 'petty_cash':
            pc_entry = PettyCash(
                date=date,
                description=petty_cash_desc or f'Customer Payment - {customer.name}',
                category=petty_cash_category or 'Customer Payment',
                type='Income',
                amount=amount
            )
            db.session.add(pc_entry)
        
        db.session.commit()

        # Recalculate everything
        recalculate_customer_totals(customer)
        db.session.commit()

        trigger_excel_sync()
        trigger_sync()
        backup_to_telegram("Added Transaction for Customer: " + customer.name)
        flash('Transaction Added!', 'success')
        return redirect(url_for('main.manage_transactions', customer_id=customer_id))
        
    transactions = Transaction.query.filter_by(customer_id=customer_id).all()
    # Fetch active installments for this customer
    customer_installments = CustomerInstallment.query.filter_by(customer_id=customer_id).all()
    banks = Bank.query.filter_by(status='Active').all()
    # Seed categories if empty
    if PettyCashCategory.query.count() == 0:
        defaults = ["Bank", "Customer Payment", "Entertainment", "Food", "Maintenance",
                    "Marketing", "Office", "Repair", "Salary", "Stationery", "Transport",
                    "Travel", "Utility"]
        for d in defaults:
            db.session.add(PettyCashCategory(name=d))
        db.session.commit()
    petty_cash_categories = [c.name for c in PettyCashCategory.query.order_by(PettyCashCategory.name).all()]

    return render_template('customer_transactions.html', 
                         customer=customer, 
                         transactions=transactions, 
                         customer_installments=customer_installments,
                         banks=banks,
                         petty_cash_categories=petty_cash_categories)

@main.route('/delete_transaction/<int:id>', methods=['GET', 'POST'])
def delete_transaction(id):
    if not verify_password():
        flash('Invalid Admin Password!', 'danger')
        tx = Transaction.query.get(id)
        if tx:
             return redirect(url_for('main.manage_transactions', customer_id=tx.customer_id))
        return redirect(url_for('main.index'))

    tx = Transaction.query.get_or_404(id)
    customer_id = tx.customer_id
    customer = Customer.query.get(customer_id)
    
    # Revert Installment if associated
    if tx.customer_installment_id:
        ci = CustomerInstallment.query.get(tx.customer_installment_id)
        if ci:
            ci.paid_amount -= tx.amount
            ci.due_amount = ci.total_amount - ci.paid_amount
    
    db.session.delete(tx)
    db.session.commit()
    
    # Final Recalculation
    recalculate_customer_totals(customer)
    db.session.commit()
    
    trigger_excel_sync()
    trigger_sync()
    backup_to_telegram("Deleted Transaction for: " + customer.name)
    flash('Transaction Deleted!', 'warning')
    return redirect(url_for('main.manage_transactions', customer_id=customer_id))

@main.route('/transaction/edit/<int:id>', methods=['POST'])
def edit_transaction_details(id):
    if not verify_password():
        flash('Invalid Admin Password!', 'danger')
        # We need customer_id to redirect
        tx = Transaction.query.get(id)
        if tx:
            # Revert to manage page if possible
            return redirect(url_for('main.manage_transactions', customer_id=tx.customer_id))
        return redirect(url_for('main.index'))

    tx = Transaction.query.get_or_404(id)
    customer = Customer.query.get(tx.customer_id)
    
    # 1. Revert Old Amount from Customer Totals
    customer.total_paid -= tx.amount
    
    # 2. Update Transaction Data
    tx.date = request.form.get('date')
    tx.amount = float(request.form.get('amount') or 0)
    tx.installment_type = request.form.get('installment_type')
    tx.bank_name = request.form.get('bank_name')
    tx.transaction_id = request.form.get('transaction_id')
    tx.remarks = request.form.get('remarks')
    
    # 3. Handle New Images (Append)
    if 'evidence' in request.files:
        files = request.files.getlist('evidence')
        new_images = []
        for file in files:
            if file and file.filename != '':
                from image_utils import save_as_webp
                filename = save_as_webp(file, current_app.config['UPLOAD_FOLDER'])
                if filename:
                    new_images.append(filename)
        
        if new_images:
            current_images = tx.images.split(',') if tx.images else []
            updated_images = current_images + new_images
            tx.images = ','.join(updated_images)
            
    # 4. Apply New Amount to Customer Totals
    customer.total_paid += tx.amount
    customer.due_amount = customer.total_price - customer.total_paid
    
    db.session.commit()
    trigger_excel_sync()
    trigger_sync()
    backup_to_telegram("Edited Transaction")
    flash('Transaction Updated!', 'success')
    return redirect(url_for('main.manage_transactions', customer_id=customer.id))


# --- Petty Cash Routes ---
@main.route('/petty_cash', methods=['GET', 'POST'])
def manage_petty_cash():
    if request.method == 'POST':
        if not verify_password():
            flash('Invalid Admin Password!', 'danger')
            return redirect(url_for('main.manage_petty_cash'))

        date = request.form.get('date')
        description = request.form.get('description')
        category = request.form.get('category')
        type = request.form.get('type')
        amount = float(request.form.get('amount') or 0)
        
        # Handle Evidence Files
        images = []
        if 'evidence' in request.files:
            files = request.files.getlist('evidence')
            for file in files:
                if file and file.filename != '':
                    from image_utils import save_as_webp
                    filename = save_as_webp(file, current_app.config['UPLOAD_FOLDER'])
                    if filename:
                        images.append(filename)
        
        new_entry = PettyCash(
            date=date,
            description=description,
            category=category,
            type=type,
            amount=amount,
            images=','.join(images)
        )
        db.session.add(new_entry)
        db.session.commit()
        trigger_excel_sync()
        trigger_sync()
        backup_to_telegram("Added Petty Cash: " + description)
        flash('Petty Cash Entry Added!', 'success')
        return redirect(url_for('main.manage_petty_cash'))
        
    # Filter Logic
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    filter_category = request.args.get('category')
    
    query = PettyCash.query
    
    if start_date:
        query = query.filter(PettyCash.date >= start_date)
    if end_date:
        query = query.filter(PettyCash.date <= end_date)
    if filter_category and filter_category != 'All':
        query = query.filter(PettyCash.category == filter_category)
        
    entries = query.order_by(PettyCash.date.desc()).all()

    # Use managed PettyCashCategory table; seed defaults if empty
    if PettyCashCategory.query.count() == 0:
        defaults = ["Bank", "Customer Payment", "Entertainment", "Food", "Maintenance",
                    "Marketing", "Office", "Repair", "Salary", "Stationery", "Transport",
                    "Travel", "Utility"]
        for d in defaults:
            db.session.add(PettyCashCategory(name=d))
        db.session.commit()

    available_categories = [c.name for c in PettyCashCategory.query.order_by(PettyCashCategory.name).all()]

    total_income = sum(e.amount for e in entries if e.type == 'Income')
    total_expense = sum(e.amount for e in entries if e.type == 'Expense')
    current_balance = total_income - total_expense
    
    pc_categories = PettyCashCategory.query.order_by(PettyCashCategory.name).all()
    return render_template('petty_cash.html', entries=entries, 
                         total_income=total_income, 
                         total_expense=total_expense, 
                         current_balance=current_balance,
                         available_categories=available_categories,
                         pc_categories=pc_categories)

# --- Petty Cash Category CRUD ---
@main.route('/petty_cash/category/add', methods=['POST'])
def add_petty_cash_category():
    name = request.form.get('name', '').strip()
    if name:
        existing = PettyCashCategory.query.filter_by(name=name).first()
        if not existing:
            db.session.add(PettyCashCategory(name=name))
            db.session.commit()
            flash(f'Category "{name}" added!', 'success')
        else:
            flash(f'Category "{name}" already exists.', 'warning')
    return redirect(url_for('main.manage_petty_cash'))

@main.route('/petty_cash/category/edit/<int:id>', methods=['POST'])
def edit_petty_cash_category(id):
    cat = PettyCashCategory.query.get_or_404(id)
    new_name = request.form.get('name', '').strip()
    if new_name and new_name != cat.name:
        old_name = cat.name
        cat.name = new_name
        # Also update all existing PettyCash entries with this category
        PettyCash.query.filter_by(category=old_name).update({'category': new_name})
        db.session.commit()
        flash(f'Category renamed to "{new_name}".', 'success')
    return redirect(url_for('main.manage_petty_cash'))

@main.route('/petty_cash/category/delete/<int:id>', methods=['POST'])
def delete_petty_cash_category(id):
    cat = PettyCashCategory.query.get_or_404(id)
    name = cat.name
    db.session.delete(cat)
    db.session.commit()
    flash(f'Category "{name}" deleted.', 'info')
    return redirect(url_for('main.manage_petty_cash'))

@main.route('/petty_cash/categories', methods=['GET'])
def get_petty_cash_categories():
    """JSON endpoint for category list (used by customer_transactions form)."""
    cats = [c.name for c in PettyCashCategory.query.order_by(PettyCashCategory.name).all()]
    from flask import jsonify
    return jsonify(cats)

@main.route('/petty_cash/export')
def export_petty_cash_report():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    filter_category = request.args.get('category')
    
    query = PettyCash.query
    
    if start_date:
        query = query.filter(PettyCash.date >= start_date)
    if end_date:
        query = query.filter(PettyCash.date <= end_date)
    if filter_category and filter_category != 'All':
        query = query.filter(PettyCash.category == filter_category)
        
    entries = query.order_by(PettyCash.date.desc()).all()
    
    try:
        data = []
        for e in entries:
            data.append({
                'Date': e.date,
                'Description': e.description,
                'Category': e.category,
                'Type': e.type,
                'Income': e.amount if e.type == 'Income' else 0,
                'Expense': e.amount if e.type == 'Expense' else 0,
                'Images': e.images
            })
            
        df = pd.DataFrame(data)
        
        # Calculate totals for the report
        total_income = df['Income'].sum() if not df.empty else 0
        total_expense = df['Expense'].sum() if not df.empty else 0
        
        # Append Total Row
        if not df.empty:
            df.loc['Total'] = pd.Series(dtype='float64')
            df.at['Total', 'Description'] = 'TOTAL'
            df.at['Total', 'Income'] = total_income
            df.at['Total', 'Expense'] = total_expense
            df.at['Total', 'Category'] = f'Balance: {total_income - total_expense}'
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Petty_Cash_Report', index=False)
            # Polish width
            worksheet = writer.sheets['Petty_Cash_Report']
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

        # Get company info
        company_name = 'NEXUS RIVER VIEW'
        data_dir = current_app.config.get('DATA_FOLDER', '.')
        settings_path = os.path.join(data_dir, 'company_settings.json')
        if os.path.exists(settings_path):
            try:
                with open(settings_path, 'r') as f:
                    s = json.load(f)
                    company_name = s.get('company_name', company_name)
            except Exception: pass
        
        safe_company_name = "".join(c for c in company_name if c.isalnum() or c in (' ', '_', '-')).strip().replace(' ', '_')
        filename = f"{safe_company_name}_Petty_Cash_Report.xlsx"
        
        output.seek(0)
        return send_file(output, download_name=filename, as_attachment=True)
    except Exception as e:
        from telegram_utils import log_debug
        log_debug(f"Excel Export Error (Petty Cash): {e}")
        import traceback
        log_debug(traceback.format_exc())
        flash(f"Error exporting Excel: {e}", "danger")
        return redirect(url_for('main.manage_petty_cash'))

@main.route('/delete_petty_cash/<int:id>', methods=['POST'])
def delete_petty_cash(id):
    if not verify_password():
        flash('Invalid Admin Password!', 'danger')
        return redirect(url_for('main.manage_petty_cash'))

    entry = PettyCash.query.get_or_404(id)
    db.session.delete(entry)
    db.session.commit()
    trigger_excel_sync()
    trigger_sync()
    backup_to_telegram("Deleted Petty Cash")
    flash('Entry Deleted!', 'info')
    return redirect(url_for('main.manage_petty_cash'))

@main.route('/petty_cash/invoice/<int:id>')
def invoice_view(id):
    entry = PettyCash.query.get_or_404(id)
    return render_template('invoice.html', entry=entry, now=datetime.now())

@main.route('/petty_cash/edit/<int:id>', methods=['POST'])
def edit_petty_cash(id):
    if not verify_password():
        flash('Invalid Admin Password!', 'danger')
        return redirect(url_for('main.manage_petty_cash'))

    entry = PettyCash.query.get_or_404(id)
    
    entry.date = request.form.get('date')
    entry.description = request.form.get('description')
    entry.category = request.form.get('category')
    entry.type = request.form.get('type')
    entry.amount = float(request.form.get('amount') or 0)
    
    # Handle New Evidence Files (Append to existing?)
    # For simplicity, if new files are uploaded, we append them.
    if 'evidence' in request.files:
        files = request.files.getlist('evidence')
        new_images = []
        for file in files:
            if file and file.filename != '':
                filename = secure_filename(file.filename)
                file.save(os.path.join(current_app.config['UPLOAD_FOLDER'], filename))
                new_images.append(filename)
        
        if new_images:
            current_images = entry.images.split(',') if entry.images else []
            updated_images = current_images + new_images
            entry.images = ','.join(updated_images)
            
    db.session.commit()
    trigger_excel_sync()
    trigger_sync()
    backup_to_telegram("Edited Petty Cash")
    flash('Entry Updated Successfully!', 'success')
    return redirect(url_for('main.manage_petty_cash'))

@main.route('/report/download')
def download_report():
    # Generate a summary report
    # Group by Director, Sum Collection (Total Paid), Outstanding (Due Amount)
    
    directors = Director.query.all()
    report_data = []
    
    try:
        grand_total_collection = 0
        grand_total_due = 0
        
        for d in directors:
            d_collection = sum(c.total_paid for c in d.customers)
            d_due = sum(c.due_amount for c in d.customers)
            
            grand_total_collection += d_collection
            grand_total_due += d_due
            
            report_data.append({
                'Director': d.name,
                'Total Collection': d_collection,
                'Total Outstanding Dues': d_due
            })
            
        # Append Grand Total
        report_data.append({
            'Director': 'GRAND TOTAL',
            'Total Collection': grand_total_collection,
            'Total Outstanding Dues': grand_total_due
        })
        
        df = pd.DataFrame(report_data)
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Summary_Report', index=False)
            
            # Polish width
            worksheet = writer.sheets['Summary_Report']
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

        # Get company info
        company_name = 'NEXUS RIVER VIEW'
        data_dir = current_app.config.get('DATA_FOLDER', '.')
        settings_path = os.path.join(data_dir, 'company_settings.json')
        if os.path.exists(settings_path):
            try:
                with open(settings_path, 'r') as f:
                    s = json.load(f)
                    company_name = s.get('company_name', company_name)
            except Exception: pass
            
        safe_company_name = "".join(c for c in company_name if c.isalnum() or c in (' ', '_', '-')).strip().replace(' ', '_')
        filename = f"{safe_company_name}_Summary_Report.xlsx"

        output.seek(0)
        
        return send_file(output, download_name=filename, as_attachment=True)
    except Exception as e:
        from telegram_utils import log_debug
        log_debug(f"Excel Export Error (Summary Report): {e}")
        import traceback
        log_debug(traceback.format_exc())
        flash(f"Error exporting Excel: {e}", "danger")
        return redirect(url_for('main.index'))

# --- Excel Helper ---
def format_excel_width(writer, sheet_name):
    """Helper to auto-adjust column width for Excel sheets."""
    if sheet_name in writer.sheets:
        worksheet = writer.sheets[sheet_name]
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

# --- New Customer Reports ---

def save_and_open_excel(output_bytes, filename):
    """
    Saves an Excel BytesIO to the user's Documents/NexusRiverView folder
    and opens it automatically. This works inside pywebview where send_file downloads don't work.
    Returns (saved_path, error_msg).
    """
    import os, subprocess
    try:
        docs_dir = os.path.join(os.path.expanduser('~'), 'Documents', 'NexusRiverView')
        os.makedirs(docs_dir, exist_ok=True)
        save_path = os.path.join(docs_dir, filename)
        with open(save_path, 'wb') as f:
            f.write(output_bytes.read())
        os.startfile(save_path)  # Opens with Excel
        return save_path, None
    except Exception as e:
        return None, str(e)

@main.route('/report/customers/all')
def download_all_customers_report():
    from telegram_utils import log_debug
    try:
        customers = Customer.query.join(Director).order_by(Director.name, Customer.customer_id).all()
        if not customers:
            log_debug("WARNING: Customer.query.join(Director) returned 0 customers. Checking without join...")
            customers = Customer.query.order_by(Customer.customer_id).all()
        
        log_debug(f"Export All: Found {len(customers)} customers.")
        
        # Check total counts
        all_c = Customer.query.count()
        all_d = Director.query.count()
        log_debug(f"Total in DB: Customers={all_c}, Directors={all_d}")
        installments = Installment.query.order_by(Installment.id).all()
        inst_names = [inst.name for inst in installments]
        data = []
        for c in customers:
            # Build installment lookup
            ci_map = {ci.installment_id: ci for ci in c.installments}
            row = {
                'Director Name': c.director.name if c.director else '',
                'Customer ID': c.customer_id,
                'Customer Name': c.name,
                'Phone': c.phone,
                'Father Name': c.father_name,
                'Mother Name': c.mother_name,
                'DOB': c.dob,
                'Religion': c.religion,
                'Profession': c.profession,
                'NID No': c.nid_no,
                'Present Address': c.present_address,
                'Permanent Address': c.permanent_address,
                'Plot No': c.plot_no,
                'Shares': int(c.shares) if c.shares == int(c.shares) else c.shares,
            }
            # Add per-installment columns
            for inst in installments:
                ci = ci_map.get(inst.id)
                row[f'{inst.name} (Total)'] = ci.total_amount if ci else 0
                row[f'{inst.name} (Paid)'] = ci.paid_amount if ci else 0
                row[f'{inst.name} (Due)'] = ci.due_amount if ci else 0
            row['Total Price'] = c.total_price
            row['Total Paid'] = c.total_paid
            row['Due Amount'] = c.due_amount
            data.append(row)

        df = pd.DataFrame(data)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='All_Customers', index=False)
            format_excel_width(writer, 'All_Customers')
        output.seek(0)
        
        # Get company info
        company_name = 'NEXUS RIVER VIEW'
        data_dir = current_app.config.get('DATA_FOLDER', '.')
        settings_path = os.path.join(data_dir, 'company_settings.json')
        if os.path.exists(settings_path):
            try:
                with open(settings_path, 'r') as f:
                    s = json.load(f)
                    company_name = s.get('company_name', company_name)
            except Exception: pass
        
        safe_company_name = "".join(c for c in company_name if c.isalnum() or c in (' ', '_', '-')).strip().replace(' ', '_')
        filename = f"{safe_company_name}_All_Customers_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        path, err = save_and_open_excel(output, filename)
        if err:
            flash(f"Export failed: {err}", "danger")
        else:
            flash(f"Exported! File saved to Documents/NexusRiverView/{filename}", "success")
        return redirect(url_for('main.index'))
    except Exception as e:
        from telegram_utils import log_debug
        log_debug(f"Excel Export Error (All Customers): {e}")
        import traceback
        log_debug(traceback.format_exc())
        flash(f"Error exporting Excel: {e}", "danger")
        return redirect(url_for('main.index'))

@main.route('/report/director/<int:id>/customers')
def download_director_customers_report(id):
    director = Director.query.get_or_404(id)
    customers = director.customers
    
    try:
        installments = Installment.query.order_by(Installment.id).all()
        data = []
        for c in customers:
            ci_map = {ci.installment_id: ci for ci in c.installments}
            row = {
                'Customer ID': c.customer_id,
                'Customer Name': c.name,
                'Phone': c.phone,
                'Father Name': c.father_name,
                'Mother Name': c.mother_name,
                'DOB': c.dob,
                'Religion': c.religion,
                'Profession': c.profession,
                'NID No': c.nid_no,
                'Present Address': c.present_address,
                'Permanent Address': c.permanent_address,
                'Plot No': c.plot_no,
                'Shares': int(c.shares) if c.shares == int(c.shares) else c.shares,
            }
            for inst in installments:
                ci = ci_map.get(inst.id)
                row[f'{inst.name} (Total)'] = ci.total_amount if ci else 0
                row[f'{inst.name} (Paid)'] = ci.paid_amount if ci else 0
                row[f'{inst.name} (Due)'] = ci.due_amount if ci else 0
            row['Total Price'] = c.total_price
            row['Total Paid'] = c.total_paid
            row['Due Amount'] = c.due_amount
            data.append(row)
            
        df = pd.DataFrame(data)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Customers', index=False)
            format_excel_width(writer, 'Customers')
        output.seek(0)
        
        # Get company info
        company_name = 'NEXUS RIVER VIEW'
        data_dir = current_app.config.get('DATA_FOLDER', '.')
        settings_path = os.path.join(data_dir, 'company_settings.json')
        if os.path.exists(settings_path):
            try:
                with open(settings_path, 'r') as f:
                    s = json.load(f)
                    company_name = s.get('company_name', company_name)
            except Exception: pass
            
        safe_company_name = "".join(c for c in company_name if c.isalnum() or c in (' ', '_', '-')).strip().replace(' ', '_')
        filename = f"{safe_company_name}_Director_{secure_filename(director.name)}_Customers_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        path, err = save_and_open_excel(output, filename)
        if err:
            flash(f"Export failed: {err}", "danger")
        else:
            flash(f"Exported! File saved to Documents/NexusRiverView/{filename}", "success")
        return redirect(url_for('main.index'))
    except Exception as e:
        from telegram_utils import log_debug
        log_debug(f"Excel Export Error (Director Customers): {e}")
        import traceback
        log_debug(traceback.format_exc())
        flash(f"Error exporting Excel: {e}", "danger")
        return redirect(url_for('main.index'))

@main.route('/report/customer/<int:id>')
def download_individual_customer_report(id):
    c = Customer.query.get_or_404(id)
    
    # 1. Profile Data
    profile_data = [{
        'Field': 'Customer ID', 'Value': c.customer_id},
        {'Field': 'Name', 'Value': c.name},
        {'Field': 'Director', 'Value': c.director.name},
        {'Field': 'Phone', 'Value': c.phone},
        {'Field': 'Father Name', 'Value': c.father_name},
        {'Field': 'Mother Name', 'Value': c.mother_name},
        {'Field': 'DOB', 'Value': c.dob},
        {'Field': 'Religion', 'Value': c.religion},
        {'Field': 'Profession', 'Value': c.profession},
        {'Field': 'NID No', 'Value': c.nid_no},
        {'Field': 'Present Address', 'Value': c.present_address},
        {'Field': 'Permanent Address', 'Value': c.permanent_address},
        {'Field': 'Plot No', 'Value': c.plot_no},
        {'Field': 'Total Price', 'Value': c.total_price},
        {'Field': 'Down Payment', 'Value': c.down_payment},
        {'Field': 'Monthly Installment', 'Value': c.monthly_installment},
        {'Field': 'Total Paid', 'Value': c.total_paid},
        {'Field': 'Due Amount', 'Value': c.due_amount}
    ]
    df_profile = pd.DataFrame(profile_data)
    
    # 2. Transactions Data
    tx_data = []
    for tx in c.transactions:
        tx_data.append({
            'Date': tx.date,
            'Amount': tx.amount,
            'Installment Type': tx.installment_type,
            'Bank Name': tx.bank_name,
            'Transaction ID': tx.transaction_id,
            'Remarks': tx.remarks
        })
    df_tx = pd.DataFrame(tx_data)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_profile.to_excel(writer, sheet_name='Profile', index=False)
        format_excel_width(writer, 'Profile')
        df_tx.to_excel(writer, sheet_name='Transactions', index=False)
        format_excel_width(writer, 'Transactions')
    output.seek(0)
    
    # Get company info
    company_name = 'NEXUS RIVER VIEW'
    data_dir = current_app.config.get('DATA_FOLDER', '.')
    settings_path = os.path.join(data_dir, 'company_settings.json')
    if os.path.exists(settings_path):
        try:
            with open(settings_path, 'r') as f:
                s = json.load(f)
                company_name = s.get('company_name', company_name)
        except Exception: pass
        
    safe_company_name = "".join(c for c in company_name if c.isalnum() or c in (' ', '_', '-')).strip().replace(' ', '_')
    filename = f"{safe_company_name}_Customer_{secure_filename(c.name)}_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    
    path, err = save_and_open_excel(output, filename)
    if err:
        flash(f"Export failed: {err}", "danger")
    else:
        flash(f"Exported! File saved to Documents/NexusRiverView/{filename}", "success")
    return redirect(url_for('main.manage_transactions', customer_id=c.id))

# --- Bank Management Routes ---
@main.route('/banks', methods=['GET', 'POST'])
def manage_banks():
    if request.method == 'POST':
        if not verify_password():
            flash('Invalid Admin Password!', 'danger')
            return redirect(url_for('main.manage_banks'))

        # Add New Bank
        new_bank = Bank(
            bank_name=request.form.get('bank_name'),
            branch=request.form.get('branch'),
            account_holder_name=request.form.get('account_holder_name'),
            joint_name=request.form.get('joint_name'),
            fhp=request.form.get('fhp'),
            address=request.form.get('address'),
            city=request.form.get('city'),
            phone=request.form.get('phone'),
            customer_id=request.form.get('customer_id'),
            account_no=request.form.get('account_no'),
            prev_account_no=request.form.get('prev_account_no'),
            account_type=request.form.get('account_type'),
            currency=request.form.get('currency'),
            status=request.form.get('status')
        )
        db.session.add(new_bank)
        db.session.commit()
        trigger_excel_sync()
        backup_to_telegram("Added Bank: " + request.form.get('bank_name'))
        flash('Bank Account Added!', 'success')
        return redirect(url_for('main.manage_banks'))
        
    banks = Bank.query.all()
    return render_template('banks.html', banks=banks)

@main.route('/bank/edit/<int:id>', methods=['POST'])
def edit_bank(id):
    if not verify_password():
        flash('Invalid Admin Password!', 'danger')
        return redirect(url_for('main.manage_banks'))

    bank = Bank.query.get_or_404(id)
    
    bank.bank_name = request.form.get('bank_name')
    bank.branch = request.form.get('branch')
    bank.account_holder_name = request.form.get('account_holder_name')
    bank.joint_name = request.form.get('joint_name')
    bank.fhp = request.form.get('fhp')
    bank.address = request.form.get('address')
    bank.city = request.form.get('city')
    bank.phone = request.form.get('phone')
    bank.customer_id = request.form.get('customer_id')
    bank.account_no = request.form.get('account_no')
    bank.prev_account_no = request.form.get('prev_account_no')
    bank.account_type = request.form.get('account_type')
    bank.currency = request.form.get('currency')
    bank.status = request.form.get('status')
    
    db.session.commit()
    trigger_excel_sync()
    backup_to_telegram("Edited Bank: " + bank.bank_name)
    flash('Bank Account Updated!', 'success')
    return redirect(url_for('main.manage_banks'))

@main.route('/bank/delete/<int:id>', methods=['POST'])
def delete_bank(id):
    if not verify_password():
        flash('Invalid Admin Password!', 'danger')
        return redirect(url_for('main.manage_banks'))

    bank = Bank.query.get_or_404(id)
    db.session.delete(bank)
    db.session.commit()
    trigger_excel_sync()
    backup_to_telegram("Deleted Bank: " + bank.bank_name)
    flash('Bank Account Deleted!', 'warning')
    return redirect(url_for('main.manage_banks'))

@main.route('/bank/<int:id>/ledger', methods=['GET', 'POST'])
def bank_ledger(id):
    bank = Bank.query.get_or_404(id)
    
    # Force recompute on view to ensure existing data is corrected
    recompute_bank_balances(id)
    
    if request.method == 'POST':
        if not verify_password():
            flash('Invalid Admin Password!', 'danger')
            return redirect(url_for('main.bank_ledger', id=id))

        date = request.form.get('date')
        cheque_no = request.form.get('cheque_no')
        ref_no = request.form.get('ref_no')
        narration = request.form.get('narration')
        transaction_details = request.form.get('transaction_details')
        category = request.form.get('category')
        tx_type = request.form.get('tx_type')
        if tx_type == 'credit':
            credit = float(request.form.get('credit') or 0)
            debit = 0.0
        else:
            debit = float(request.form.get('debit') or 0)
            credit = 0.0
        
        # Convert YYYY-MM-DD input to DD-MM-YYYY storage
        try:
            date_obj = datetime.strptime(date, '%Y-%m-%d')
            date = date_obj.strftime('%d-%m-%Y')
        except ValueError:
             # Already in format or invalid? keep as is
             pass

        new_tx = BankTransaction(
            date=date,
            cheque_no=cheque_no,
            ref_no=ref_no,
            narration=narration,
            transaction_details=transaction_details,
            category=category,
            debit=debit,
            credit=credit,
            balance=0,
            bank=bank
        )

        db.session.add(new_tx)
        db.session.commit()
        
        # Recompute Balances
        recompute_bank_balances(id)
        
        trigger_excel_sync()
        backup_to_telegram("Added Bank Ledger Tx")
        flash('Transaction Added to Ledger!', 'success')
        return redirect(url_for('main.bank_ledger', id=id))
        
    transactions = BankTransaction.query.filter_by(bank_id=id).all()
    
    # Sort in Python to handle Date Parsing correctly
    # DB Date is String, format DD-MM-YYYY preferred
    def parse_tx_date(tx):
        for fmt in ('%d-%m-%Y', '%Y-%m-%d'):
            try:
                return datetime.strptime(tx.date, fmt)
            except ValueError:
                pass
        return datetime.min # Fallback

    transactions.sort(key=lambda x: (parse_tx_date(x), x.id), reverse=True)
    
    # --- Date Filter Logic ---
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    
    if start_date_str or end_date_str:
        filtered_transactions = []
        
        # Convert filter inputs (YYYY-MM-DD -> datetime)
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d') if start_date_str else None
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d') if end_date_str else None
        
        for tx in transactions:
            try:
                # Use helper or direct parse
                tx_date = parse_tx_date(tx)
                
                # Apply Filter
                if start_date and tx_date < start_date:
                    continue
                if end_date and tx_date > end_date:
                    continue
                
                filtered_transactions.append(tx)
            except ValueError:
                filtered_transactions.append(tx)
                
        transactions = filtered_transactions


    categories = PettyCashCategory.query.order_by(PettyCashCategory.name).all()
    return render_template('bank_ledger.html', bank=bank, transactions=transactions, 
                           start_date=start_date_str, end_date=end_date_str, categories=categories)

def recompute_bank_balances(bank_id):
    """
    Recalculates the running balance for all transactions of a specific bank.
    Sorts by Date (asc) and then ID (asc).
    Handles mixed date formats (DD-MM-YYYY vs YYYY-MM-DD).
    """
    transactions = BankTransaction.query.filter_by(bank_id=bank_id).all()
    
    def parse_date(date_str):
        # Prioritize DD-MM-YYYY
        for fmt in ('%d-%m-%Y', '%Y-%m-%d'):
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                pass
        return datetime.min

    # Sort transactions (Oldest First for Calculation)
    transactions.sort(key=lambda x: (parse_date(x.date), x.id))
    
    running_balance = 0.0
    for tx in transactions:
        running_balance += (tx.credit - tx.debit)
        tx.balance = running_balance
        
    db.session.commit()

@main.route('/bank/transaction/delete/<int:id>', methods=['POST'])
def delete_bank_transaction(id):
    if not verify_password():
        flash('Invalid Admin Password!', 'danger')
        # We need bank_id to redirect
        tx = BankTransaction.query.get(id)
        if tx:
            return redirect(url_for('main.bank_ledger', id=tx.bank_id))
        return redirect(url_for('main.manage_banks'))

    tx = BankTransaction.query.get_or_404(id)
    bank_id = tx.bank_id
    
    db.session.delete(tx)
    db.session.commit()
    
    recompute_bank_balances(bank_id)
    
    trigger_excel_sync()
    backup_to_telegram("Deleted Bank Tx: " + (tx.narration or str(id)))
    flash('Transaction Deleted & Balances Recomputed!', 'warning')
    return redirect(url_for('main.bank_ledger', id=bank_id))

@main.route('/bank/transaction/edit/<int:id>', methods=['POST'])
def edit_bank_transaction(id):
    if not verify_password():
        flash('Invalid Admin Password!', 'danger')
        tx = BankTransaction.query.get(id)
        if tx:
            return redirect(url_for('main.bank_ledger', id=tx.bank_id))
        return redirect(url_for('main.manage_banks'))

    tx = BankTransaction.query.get_or_404(id)
    bank_id = tx.bank_id
    
    # Update Data
    tx.date = request.form.get('date')
    tx.cheque_no = request.form.get('cheque_no')
    tx.ref_no = request.form.get('ref_no')
    tx.narration = request.form.get('narration')
    tx.transaction_details = request.form.get('transaction_details')
    tx.category = request.form.get('category')
    tx_type = request.form.get('tx_type')
    if tx_type == 'credit':
        tx.credit = float(request.form.get('credit') or 0)
        tx.debit = 0.0
    else:
        tx.debit = float(request.form.get('debit') or 0)
        tx.credit = 0.0
    
    # Save first to establish new values
    db.session.commit()
    
    recompute_bank_balances(bank_id)
    
    trigger_excel_sync()
    backup_to_telegram("Edited Bank Tx: " + (tx.narration or str(id)))
    flash('Transaction updated successfully!', 'success')
    return redirect(url_for('main.view_bank_transactions', id=bank_id))

# --- Voucher System Routes ---

@main.route('/vouchers')
def manage_vouchers():
    v_type = request.args.get('type')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    search = request.args.get('search')
    
    query = Voucher.query
    if v_type:
        query = query.filter_by(type=v_type)
    if start_date:
        query = query.filter(Voucher.date >= start_date)
    if end_date:
        query = query.filter(Voucher.date <= end_date)
    if search:
        query = query.filter(
            (Voucher.voucher_no.contains(search)) | 
            (Voucher.description.contains(search))
        )
        
    vouchers = query.order_by(Voucher.date.desc(), Voucher.id.desc()).all()
    return render_template('vouchers.html', vouchers=vouchers, v_type=v_type, start_date=start_date, end_date=end_date, search=search)

@main.route('/vouchers/add/<v_type>', methods=['GET', 'POST'])
def add_voucher(v_type):
    if request.method == 'POST':
        voucher_no = request.form.get('voucher_no')
        date = request.form.get('date')
        party_id = request.form.get('party_id')
        customer_id = request.form.get('customer_id')
        description = request.form.get('description')
        total_amount = float(request.form.get('total_amount') or 0)
        amount_paid = float(request.form.get('amount_paid') or 0)
        category = request.form.get('category')
        payment_method = request.form.get('payment_method')
        bank_id = request.form.get('bank_id')
        notes = request.form.get('notes')
        
        # Calculations
        due_amount = total_amount - amount_paid
        payment_percentage = (amount_paid / total_amount * 100) if total_amount > 0 else 0
        
        v = Voucher(
            voucher_no=voucher_no,
            type=v_type,
            date=date,
            party_id=party_id if party_id else None,
            customer_id=customer_id if customer_id else None,
            description=description,
            total_amount=total_amount,
            amount_paid=amount_paid,
            due_amount=due_amount,
            payment_percentage=payment_percentage,
            payment_method=payment_method,
            bank_id=bank_id if bank_id else None,
            category=category,
            notes=notes,
            is_payment=request.form.get('is_payment') == 'on'
        )
        
        # Handle Attachment
        file = request.files.get('attachment')
        if file and file.filename:
            filename = secure_filename(file.filename)
            file.save(os.path.join(current_app.config['UPLOAD_FOLDER'], filename))
            v.attachment = filename
            
        db.session.add(v)
        db.session.commit()
        
        # Process Financials
        process_voucher_financials(v.id, action='add')
        
        backup_to_telegram(f"Added {v_type} Voucher: {voucher_no}")
        flash(f'{v_type} Voucher added successfully!', 'success')
        return redirect(url_for('main.manage_vouchers'))
        
    # GET request
    suggested_no = generate_voucher_number(v_type)
    parties = Party.query.order_by(Party.name).all()
    banks = Bank.query.filter_by(status='Active').all()
    categories = PettyCashCategory.query.order_by(PettyCashCategory.name).all()
    customers = Customer.query.order_by(Customer.name).all()
    today_str = datetime.now().strftime('%Y-%m-%d')
    return render_template('voucher_form.html', v_type=v_type, suggested_no=suggested_no, 
                           parties=parties, banks=banks, categories=categories, customers=customers, today_str=today_str)

@main.route('/vouchers/edit/<int:id>', methods=['GET', 'POST'])
def edit_voucher(id):
    v = Voucher.query.get_or_404(id)
    if request.method == 'POST':
        v.voucher_no = request.form.get('voucher_no')
        v.date = request.form.get('date')
        v.party_id = request.form.get('party_id') if request.form.get('party_id') else None
        v.customer_id = request.form.get('customer_id') if request.form.get('customer_id') else None
        v.description = request.form.get('description')
        v.total_amount = float(request.form.get('total_amount') or 0)
        v.amount_paid = float(request.form.get('amount_paid') or 0)
        v.category = request.form.get('category')
        v.payment_method = request.form.get('payment_method')
        v.bank_id = request.form.get('bank_id') if request.form.get('bank_id') else None
        v.notes = request.form.get('notes')
        v.is_payment = request.form.get('is_payment') == 'on'
        
        # Recalculate
        v.due_amount = v.total_amount - v.amount_paid
        v.payment_percentage = (v.amount_paid / v.total_amount * 100) if v.total_amount > 0 else 0
        
        # Handle Attachment
        file = request.files.get('attachment')
        if file and file.filename:
            from image_utils import save_as_webp
            filename = save_as_webp(file, current_app.config['UPLOAD_FOLDER'])
            if filename:
                v.attachment = filename
                db.session.commit()
        
        # Re-process Financials
        process_voucher_financials(v.id, action='edit')
        
        backup_to_telegram(f"Edited Voucher: {v.voucher_no}")
        flash('Voucher updated successfully!', 'success')
        return redirect(url_for('main.manage_vouchers'))
        
    parties = Party.query.order_by(Party.name).all()
    banks = Bank.query.filter_by(status='Active').all()
    categories = PettyCashCategory.query.order_by(PettyCashCategory.name).all()
    customers = Customer.query.order_by(Customer.name).all()
    return render_template('voucher_form.html', v=v, v_type=v.type, 
                           parties=parties, banks=banks, categories=categories, customers=customers)

@main.route('/vouchers/delete/<int:id>', methods=['POST'])
def delete_voucher(id):
    password = request.form.get('admin_password')
    # Use existing validation logic or simple check
    if password != current_app.config.get('ADMIN_PASSWORD', 'admin123'):
        flash("Invalid admin password.", "danger")
        return redirect(url_for('main.manage_vouchers'))
        
    v = Voucher.query.get_or_404(id)
    v_no = v.voucher_no
    
    # Process financials before deleting object
    process_voucher_financials(v.id, action='delete')
    
    db.session.delete(v)
    db.session.commit()
    
    backup_to_telegram(f"Deleted Voucher: {v_no}")
    flash('Voucher deleted successfully!', 'info')
    return redirect(url_for('main.manage_vouchers'))

@main.route('/vouchers/print/<int:id>')
def print_voucher(id):
    v = Voucher.query.get_or_404(id)
    return render_template('voucher_print.html', v=v, datetime=datetime)

@main.route('/contra')
def manage_contra_entries():
    contra_entries = ContraEntry.query.order_by(ContraEntry.date.desc(), ContraEntry.id.desc()).all()
    return render_template('contra_list.html', contra_entries=contra_entries)

@main.route('/contra/add', methods=['GET', 'POST'])
def add_contra_entry():
    if request.method == 'POST':
        contra_no = request.form.get('contra_no')
        date = request.form.get('date')
        from_acc = request.form.get('from_account')
        to_acc = request.form.get('to_account')
        bank_id = request.form.get('bank_id')
        cheque_no = request.form.get('cheque_no')
        amount = float(request.form.get('amount') or 0)
        description = request.form.get('description')
        
        # Handle file uploads
        uploaded_files = []
        files = request.files.getlist('attachments')
        if files:
            upload_folder = current_app.config['UPLOAD_FOLDER']
            for file in files:
                if file and file.filename:
                    from image_utils import save_as_webp
                    filename = save_as_webp(file, upload_folder)
                    if filename:
                        uploaded_files.append(filename)
        
        c = ContraEntry(
            contra_no=contra_no,
            date=date,
            from_account=from_acc,
            to_account=to_acc,
            bank_id=bank_id if bank_id else None,
            amount=amount,
            description=description,
            cheque_no=cheque_no,
            attachments=",".join(uploaded_files) if uploaded_files else None
        )
        db.session.add(c)
        db.session.commit()
        process_contra_financials(c.id, action='add')
        
        flash('Contra entry added successfully!', 'success')
        return redirect(url_for('main.manage_contra_entries'))
        
    suggested_no = generate_contra_number()
    banks = Bank.query.filter_by(status='Active').all()
    today_str = datetime.now().strftime('%Y-%m-%d')
    return render_template('contra_form.html', suggested_no=suggested_no, banks=banks, today_str=today_str)

@main.route('/contra/edit/<int:id>', methods=['GET', 'POST'])
def edit_contra_entry(id):
    c = ContraEntry.query.get_or_404(id)
    if request.method == 'POST':
        c.contra_no = request.form.get('contra_no')
        c.date = request.form.get('date')
        c.from_account = request.form.get('from_account')
        c.to_account = request.form.get('to_account')
        c.bank_id = request.form.get('bank_id') if request.form.get('bank_id') else None
        c.amount = float(request.form.get('amount') or 0)
        c.description = request.form.get('description')
        c.cheque_no = request.form.get('cheque_no')
        
        # Handle new file uploads
        files = request.files.getlist('attachments')
        if files and any(file.filename for file in files):
            uploaded_files = c.attachments.split(',') if c.attachments else []
            upload_folder = current_app.config['UPLOAD_FOLDER']
            for file in files:
                if file and file.filename:
                    from image_utils import save_as_webp
                    filename = save_as_webp(file, upload_folder)
                    if filename:
                        uploaded_files.append(filename)
            c.attachments = ",".join([f for f in uploaded_files if f])
        
        db.session.commit()
        process_contra_financials(c.id, action='edit')
        
        flash('Contra entry updated successfully!', 'success')
        return redirect(url_for('main.manage_contra_entries'))
        
    banks = Bank.query.filter_by(status='Active').all()
    return render_template('contra_form.html', contra=c, banks=banks)

@main.route('/contra/delete/<int:id>', methods=['POST'])
def delete_contra_entry(id):
    c = ContraEntry.query.get_or_404(id)
    
    # Delete associated files
    if c.attachments:
        upload_folder = current_app.config['UPLOAD_FOLDER']
        for filename in c.attachments.split(','):
            file_path = os.path.join(upload_folder, filename)
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                except:
                    pass
                    
    process_contra_financials(c.id, action='delete')
    db.session.delete(c)
    db.session.commit()
    flash('Contra entry deleted successfully!', 'success')
    return redirect(url_for('main.manage_contra_entries'))

# --- Reports Dashboard ---

@main.route('/reports')
def reports_dashboard():
    # Summaries
    total_due_vouchers = db.session.query(db.func.sum(Voucher.due_amount)).filter(Voucher.type == 'Debit').scalar() or 0
    total_received_vouchers = db.session.query(db.func.sum(Voucher.amount_paid)).filter(Voucher.type == 'Credit').scalar() or 0
    
    # Party Summary
    parties = Party.query.all()
    party_stats = []
    for p in parties:
        due = p.current_balance # Positive means we owe them
        if abs(due) > 0:
            party_stats.append({'name': p.name, 'balance': due})
            
    # Bank Summaries
    banks = Bank.query.all()
    bank_stats = []
    for b in banks:
        # Last balance
        last_tx = BankTransaction.query.filter_by(bank_id=b.id).order_by(BankTransaction.id.desc()).first()
        bank_stats.append({'name': b.bank_name, 'balance': last_tx.balance if last_tx else 0})
        
    return render_template('reports.html', 
                          total_due_vouchers=total_due_vouchers, 
                          total_received_vouchers=total_received_vouchers,
                          party_stats=party_stats,
                          bank_stats=bank_stats)

@main.route('/reports/daily_cash')
def daily_cash_report():
    date = request.args.get('date', datetime.now().strftime("%Y-%m-%d"))
    results = get_daily_cash_report(date)
    return render_template('reports_daily_cash.html', results=results, date=date)

@main.route('/reports/daily_cash/export/excel', methods=['POST'])
def daily_cash_report_export_excel():
    """Export daily cash report to Excel."""
    import io
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from flask import send_file

    date = request.form.get('date', datetime.now().strftime('%Y-%m-%d'))
    results = get_daily_cash_report(date)

    # Load company info
    company_name = 'Company'
    company_address = ''
    data_dir = current_app.config.get('DATA_FOLDER', '.')
    settings_path = os.path.join(data_dir, 'company_settings.json')
    if os.path.exists(settings_path):
        try:
            with open(settings_path, 'r') as f:
                s = json.load(f)
                company_name = s.get('company_name', company_name)
                company_address = s.get('company_address', company_address)
        except Exception:
            pass

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Daily Cash Report'

    center = Alignment(horizontal='center', vertical='center')
    bold = Font(bold=True)
    header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
    header_font_white = Font(bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Header
    ws.merge_cells('A1:D1')
    ws['A1'] = company_name
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = center

    ws.merge_cells('A2:D2')
    ws['A2'] = company_address
    ws['A2'].alignment = center

    ws.merge_cells('A3:D3')
    ws['A3'] = f'Daily Cash Report — {date}'
    ws['A3'].font = Font(bold=True, size=13)
    ws['A3'].alignment = center

    # Balance Summary
    ws['A5'] = 'Cash in Hand (Previous Day):'
    ws['B5'] = results['prev_cash']
    ws['B5'].number_format = '#,##0.00'
    ws['A6'] = "Today's Cash Closing:"
    ws['B6'] = results['today_cash']
    ws['B6'].number_format = '#,##0.00'
    ws['C5'] = 'Bank Balance (Previous Day):'
    ws['D5'] = results['prev_bank']
    ws['D5'].number_format = '#,##0.00'
    ws['C6'] = "Today's Bank Closing:"
    ws['D6'] = results['today_bank']
    ws['D6'].number_format = '#,##0.00'

    for row in [5, 6]:
        for col in ['A', 'C']:
            ws[f'{col}{row}'].font = bold

    # Table header
    trow = 8
    for idx, h in enumerate(['Description', 'Category', 'Type', 'Amount'], start=1):
        cell = ws.cell(row=trow, column=idx, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    total_in = 0.0
    total_out = 0.0
    for r_idx, item in enumerate(results['transactions'], start=trow + 1):
        amt = item.amount
        if item.type == 'Income':
            total_in += amt
            type_label = 'Income'
            amt_color = '166534'
        else:
            total_out += amt
            type_label = 'Payment'
            amt_color = '991B1B'

        ws.cell(row=r_idx, column=1, value=item.description).border = thin_border
        ws.cell(row=r_idx, column=2, value=item.category).border = thin_border
        ws.cell(row=r_idx, column=3, value=type_label).border = thin_border
        amt_cell = ws.cell(row=r_idx, column=4, value=amt)
        amt_cell.font = Font(bold=True, color=amt_color)
        amt_cell.number_format = '#,##0.00'
        amt_cell.border = thin_border

    # Totals
    last_row = trow + 1 + len(results['transactions'])
    ws.cell(row=last_row, column=3, value='Total Cash In:').font = bold
    tc = ws.cell(row=last_row, column=4, value=total_in)
    tc.font = Font(bold=True, color='166534')
    tc.number_format = '#,##0.00'

    ws.cell(row=last_row + 1, column=3, value='Total Cash Out:').font = bold
    tco = ws.cell(row=last_row + 1, column=4, value=total_out)
    tco.font = Font(bold=True, color='991B1B')
    tco.number_format = '#,##0.00'

    ws.cell(row=last_row + 2, column=3, value='Net Cash Shift:').font = Font(bold=True)
    nc = ws.cell(row=last_row + 2, column=4, value=total_in - total_out)
    nc.font = Font(bold=True)
    nc.number_format = '#,##0.00'

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Sanitize company name for filename
    safe_company_name = "".join(c for c in company_name if c.isalnum() or c in (' ', '_', '-')).strip().replace(' ', '_')
    filename = f'{safe_company_name}_Daily_Cash_Report_{date}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )



@main.route('/reports/bank/<int:bank_id>')
def bank_report(bank_id):
    bank = Bank.query.get_or_404(bank_id)
    transactions = BankTransaction.query.filter_by(bank_id=bank_id).order_by(BankTransaction.date.desc(), BankTransaction.id.desc()).all()
    return render_template('reports_bank.html', bank=bank, transactions=transactions)

@main.route('/bank/<int:id>/export')
def export_bank_ledger(id):
    bank = Bank.query.get_or_404(id)
    # Raw query: no sorting
    transactions = BankTransaction.query.filter_by(bank_id=id).all()
    
    from sync_manager import sync_manager
    data = [sync_manager._model_to_dict('bank_transaction', tx) for tx in transactions]
        
    df = pd.DataFrame(data)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        sheet_name = 'Ledger'
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        # No formatting/auto-width as requested
        
    # Get company info
    company_name = 'NEXUS RIVER VIEW'
    data_dir = current_app.config.get('DATA_FOLDER', '.')
    settings_path = os.path.join(data_dir, 'company_settings.json')
    if os.path.exists(settings_path):
        try:
            with open(settings_path, 'r') as f:
                s = json.load(f)
                company_name = s.get('company_name', company_name)
        except Exception: pass
        
    safe_company_name = "".join(c for c in company_name if c.isalnum() or c in (' ', '_', '-')).strip().replace(' ', '_')
    safe_name = secure_filename(bank.bank_name) or "Bank"
    filename = f"{safe_company_name}_Bank_Ledger_{safe_name}.xlsx"
    
    # Optional: Save a copy to the data directory for GUI users
    try:
        data_dir = current_app.config.get('DATA_FOLDER', '.')
        disk_path = os.path.join(data_dir, filename)
        with open(disk_path, 'wb') as f:
            f.write(output.getvalue())
        from telegram_utils import log_debug
        log_debug(f"Bank Ledger saved to disk: {disk_path}")
    except Exception as e:
        print(f"Failed to save disk backup: {e}")

    output.seek(0)
    return send_file(
        output, 
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        download_name=filename, 
        as_attachment=True
    )

@main.route('/report/bank/all')
def export_all_bank_data():
    from sync_manager import sync_manager
    banks = Bank.query.all()
    bank_transactions = BankTransaction.query.all()

    banks_data = [sync_manager._model_to_dict('bank', b) for b in banks]
    tx_data = [sync_manager._model_to_dict('bank_transaction', tx) for tx in bank_transactions]

    df_banks = pd.DataFrame(banks_data)
    df_tx = pd.DataFrame(tx_data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_banks.to_excel(writer, sheet_name='Banks', index=False)
        df_tx.to_excel(writer, sheet_name='Bank_Transactions', index=False)

    # Optional: Save a copy to the data directory
    # Get company info
    company_name = 'NEXUS RIVER VIEW'
    data_dir = current_app.config.get('DATA_FOLDER', '.')
    settings_path = os.path.join(data_dir, 'company_settings.json')
    if os.path.exists(settings_path):
        try:
            with open(settings_path, 'r') as f:
                s = json.load(f)
                company_name = s.get('company_name', company_name)
        except Exception: pass
        
    safe_company_name = "".join(c for c in company_name if c.isalnum() or c in (' ', '_', '-')).strip().replace(' ', '_')
    filename = f"{safe_company_name}_All_Bank_Data_Raw.xlsx"
    try:
        data_dir = current_app.config.get('DATA_FOLDER', '.')
        disk_path = os.path.join(data_dir, filename)
        with open(disk_path, 'wb') as f:
            f.write(output.getvalue())
        from telegram_utils import log_debug
        log_debug(f"All Bank Data saved to disk: {disk_path}")
    except Exception as e:
        print(f"Failed to save disk backup: {e}")

    output.seek(0)
    return send_file(
        output, 
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        download_name=filename, 
        as_attachment=True
    )

# --- Settings & Restore ---
@main.route('/settings')
def settings():
    # Load company settings
    company_name = "Company Name"
    company_address = ""
    data_dir = current_app.config.get('DATA_FOLDER', '.')
    settings_path = os.path.join(data_dir, 'company_settings.json')
    if os.path.exists(settings_path):
        try:
            with open(settings_path, 'r') as f:
                settings_data = json.load(f)
                company_name = settings_data.get('company_name', company_name)
                company_address = settings_data.get('company_address', company_address)
        except Exception:
            pass

    # List available backups
    backups = []
    backup_dir = os.path.join(current_app.root_path, 'backups')
    if os.path.exists(backup_dir):
        files = sorted(os.listdir(backup_dir), reverse=True)
        for f in files:
            if f.endswith('.xlsx'):
                backups.append(f)
    return render_template('settings.html', backups=backups, company_name=company_name, company_address=company_address)

@main.route('/settings/company', methods=['POST'])
def save_company_settings():
    company_name = request.form.get('company_name', '')
    company_address = request.form.get('company_address', '')
    
    data_dir = current_app.config.get('DATA_FOLDER', '.')
    settings_path = os.path.join(data_dir, 'company_settings.json')
    try:
        with open(settings_path, 'w') as f:
            json.dump({'company_name': company_name, 'company_address': company_address}, f)
        flash('Company settings updated successfully!', 'success')
    except Exception as e:
        flash(f'Failed to save company settings: {e}', 'danger')
        
    return redirect(url_for('main.settings'))

@main.route('/settings/restore', methods=['POST'])
def restore_data():
    if 'backup_file' in request.files:
        file = request.files['backup_file']
        if file and file.filename != '':
            filename = secure_filename(file.filename)
            filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            
            success, msg = restore_from_excel(filepath)
            os.remove(filepath) # Clean up temp file
            
            if success:
                flash('System successfully restored from upload!', 'success')
            else:
                flash(f'Restore failed: {msg}', 'danger')
                
    elif 'backup_filename' in request.form:
        # Restore from internal backup
        filename = request.form.get('backup_filename')
        filepath = os.path.join(current_app.root_path, 'backups', filename)
        if os.path.exists(filepath):
            success, msg = restore_from_excel(filepath)
            if success:
                flash(f'System restored from {filename}', 'success')
            else:
                flash(f'Restore failed: {msg}', 'danger')
        else:
            flash('Backup file not found.', 'danger')
            
    return redirect(url_for('main.settings'))


@main.route('/sync', methods=['GET', 'POST'])
def sync_resolution():
    from sync_manager import sync_manager
    mismatch_details = current_app.config.get('SYNC_MISMATCH')
    
    if not mismatch_details:
        current_app.config['SYNC_MISMATCH'] = None
        flash('No sync mismatch details found. Proceeding to Dashboard.', 'info')
        return redirect(url_for('main.index'))

    if request.method == 'POST':
        choice = request.form.get('choice')
        if choice == 'sheet_to_db':
            success, msg = sync_manager.restore_db_from_sheets()
            if success:
                flash('Local database restored from Google Sheets!', 'success')
                current_app.config['SYNC_MISMATCH'] = None
            else:
                flash(f'Sync failed: {msg}', 'danger')
        elif choice == 'db_to_sheet':
            success, msg = sync_manager.sync_to_sheets()
            if success:
                flash('Google Sheets updated with local data!', 'success')
                current_app.config['SYNC_MISMATCH'] = None
            else:
                flash(f'Sync failed: {msg}', 'danger')
        
        return redirect(url_for('main.index'))

    return render_template('sync_resolution.html', details=mismatch_details)

    # (Consolidated into global_request_processor at the top of the file)
    pass

# --- Notifications / Logs ---
@main.route('/notifications')
def notifications():
    from telegram_utils import get_log_path
    log_file = get_log_path()
    logs = ""
    if os.path.exists(log_file):
        try:
            # Read the last 500 lines for performance
            with open(log_file, 'r', encoding='utf-8') as f:
                lines = f.readlines()
                logs = "".join(lines[-500:])
        except Exception as e:
            logs = f"Error reading logs: {e}"
    
    return render_template('notifications.html', logs=logs)

@main.route('/clear_logs', methods=['POST'])
def clear_logs():
    if not verify_password():
        flash('Invalid Admin Password!', 'danger')
        return redirect(url_for('main.notifications'))
        
    from telegram_utils import get_log_path
    log_file = get_log_path()
    try:
        if os.path.exists(log_file):
            with open(log_file, 'w', encoding='utf-8') as f:
                f.write("") # Clear
        flash('Logs cleared successfully.', 'success')
    except Exception as e:
        flash(f'Failed to clear logs: {e}', 'danger')
        
    return redirect(url_for('main.notifications'))

# --- AI Assistant ---
@main.route('/ai_assistant')
def ai_assistant():
    return render_template('ai_assistant.html')

@main.route('/favicon.ico')
def favicon():
    return send_from_directory(os.path.join(current_app.root_path, 'static'),
                               'favicon.ico', mimetype='image/vnd.microsoft.icon')

@main.route('/api/chat', methods=['POST'])
def api_ai_chat():
    from ai_chat import chat_with_db
    data = request.get_json()
    if not data or 'message' not in data:
        return {"error": "No message provided"}, 400
        
    user_message = data['message']
    history = data.get('history', [])
    
    try:
        reply = chat_with_db(user_message, chat_history=history)
        return {"reply": reply}
    except Exception as e:
        from telegram_utils import log_debug
        log_debug(f"AI Route Error: {e}")
        return {"error": str(e)}, 500

# --- Global Error Handler ---
@main.app_errorhandler(Exception)
def handle_exception(e):
    from telegram_utils import log_debug
    from flask import request, jsonify
    import traceback
    
    error_details = traceback.format_exc()
    log_debug(f"Unhandled Exception in {request.path}: {e}\n{error_details}")
    
    # If this is an API route, return JSON instead of a redirect
    if request.path.startswith('/api/'):
        return jsonify({
            'error': str(e),
            'type': type(e).__name__
        }), 500
        
    # Prevent redirect loops for 404s or background browser requests
    from werkzeug.exceptions import NotFound
    if isinstance(e, NotFound):
        return f"404 Not Found: {request.path}", 404

    # Log the specific error for debugging
    print(f"ERROR on {request.path}: {e}")
    
    # Return error page instead of redirect loop
    return render_template('error.html', error=str(e), path=request.path), 500

